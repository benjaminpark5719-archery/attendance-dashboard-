"""
=================================================================
보물섬수산 출근현황 관리 HR 시스템 v2.1
Supabase REST API / 25개 인사항목 / TBM 안전관리
=================================================================
"""
import streamlit as st
import pandas as pd
import requests
import base64
from datetime import datetime, date
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="보물섬수산 HR", page_icon="🐟", layout="wide")

# ─── Supabase REST API ───────────────────────────────────────
SB_URL = st.secrets["supabase"]["url"]
SB_KEY = st.secrets["supabase"]["key"]
HEADERS = {
    "apikey": SB_KEY,
    "Authorization": f"Bearer {SB_KEY}",
    "Content-Type": "application/json",
}

def sb_select(table, params="", order=""):
    url = f"{SB_URL}/rest/v1/{table}?select=*"
    if params: url += f"&{params}"
    if order: url += f"&order={order}"
    r = requests.get(url, headers=HEADERS)
    return r.json() if r.ok and r.text.strip() else []

def sb_insert(table, data):
    r = requests.post(f"{SB_URL}/rest/v1/{table}", json=data,
                      headers={**HEADERS, "Prefer": "return=representation"})
    return r.ok, r.text

def sb_upsert(table, data, on_conflict):
    r = requests.post(f"{SB_URL}/rest/v1/{table}?on_conflict={on_conflict}",
                      json=data,
                      headers={**HEADERS, "Prefer": "return=representation,resolution=merge-duplicates"})
    return r.ok, r.text

def sb_update(table, data, filters):
    r = requests.patch(f"{SB_URL}/rest/v1/{table}?{filters}", json=data,
                       headers={**HEADERS, "Prefer": "return=representation"})
    return r.ok, r.text

def sb_delete(table, filters):
    r = requests.delete(f"{SB_URL}/rest/v1/{table}?{filters}", headers=HEADERS)
    return r.ok

def sb_health():
    try:
        r = requests.get(f"{SB_URL}/rest/v1/departments?select=name", headers=HEADERS, timeout=5)
        return r.ok
    except: return False

def fmt_time(v):
    """시간 포맷 자동변환: 0900→09:00, 900→09:00, 09:00→09:00"""
    if not v: return ""
    v=v.strip().replace(":","").replace(" ","")
    if not v.isdigit(): return v
    if len(v)==3: v="0"+v        # 900 → 0900
    if len(v)!=4: return v
    h,m=int(v[:2]),int(v[2:])
    if h>23 or m>59: return v    # 유효성 실패시 원본 반환
    return f"{h:02d}:{m:02d}"

STATUS_OPTIONS = ["출근완료", "지각", "결근", "휴무", "조퇴"]
DEFAULT_DEPTS = ["회계부","회주방","식당 홀","식당 주방","배송팀","물류팀","해썹가공공장","마트 입점팀"]

def init_departments():
    existing = sb_select("departments")
    names = {d["name"] for d in existing}
    new = [{"name": n} for n in DEFAULT_DEPTS if n not in names]
    if new: sb_insert("departments", new)

if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False

# ─── 엑셀 다운로드 ──────────────────────────────────────────
def to_excel(df, sheet='데이터'):
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        df.to_excel(w, sheet_name=sheet, index=False)
        ws = w.sheets[sheet]
        hf = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        hfn = Font(bold=True, color="FFFFFF", size=11)
        tb = Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
        for c in ws[1]: c.fill=hf; c.font=hfn; c.alignment=Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2,max_row=len(df)+1):
            for c in row: c.border=tb; c.alignment=Alignment(horizontal="center")
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(ml+4, 25)
    out.seek(0)
    return out.getvalue()

# ─── 로그인 ──────────────────────────────────────────────────
def show_login():
    st.markdown("<div style='text-align:center;margin-top:3rem'>"
                "<h1>🐟 보물섬수산</h1>"
                "<h3 style='color:#64748b'>출근현황 관리 HR 시스템</h3></div>",
                unsafe_allow_html=True)
    c1,c2,c3 = st.columns([1,1.2,1])
    with c2:
        st.markdown("---")
        with st.form("login"):
            u = st.text_input("아이디", value="admin")
            p = st.text_input("비밀번호", type="password")
            if st.form_submit_button("로그인", use_container_width=True, type="primary"):
                if u=="admin" and p=="admin1234!":
                    st.session_state['logged_in']=True; st.rerun()
                else: st.error("아이디 또는 비밀번호가 올바르지 않습니다.")
        st.caption("초기 계정: admin / admin1234!")

# ═══════════════════════════════════════════════════════════════
#  📊 전사현황
# ═══════════════════════════════════════════════════════════════
def page_dashboard():
    st.title("📊 전사 출근 현황")
    td = st.date_input("기준일", value=date.today())
    ts = td.strftime("%Y-%m-%d")
    emps = sb_select("employees","employment_status=eq.재직",order="dept_name,emp_no")
    atts = sb_select("attendance",f"date=eq.{ts}")
    if not emps:
        st.info("등록된 직원이 없습니다. '직원관리'에서 먼저 등록하세요."); return
    am = {a["emp_no"]:a for a in atts}
    depts = sorted(set(e["dept_name"] for e in emps))
    stats=[]
    tot={"부서":"합계","총인원":0,"출근완료":0,"지각":0,"결근":0,"휴무":0,"미입력":0}
    for d in depts:
        de=[e for e in emps if e["dept_name"]==d]
        n=len(de)
        comp=sum(1 for e in de if am.get(e["emp_no"],{}).get("status")=="출근완료")
        late=sum(1 for e in de if am.get(e["emp_no"],{}).get("status")=="지각")
        abst=sum(1 for e in de if am.get(e["emp_no"],{}).get("status")=="결근")
        off=sum(1 for e in de if am.get(e["emp_no"],{}).get("status")=="휴무")
        ni=n-comp-late-abst-off
        w=n-off
        r=f"{round(comp/w*100,1)}%" if w>0 else "0%"
        row={"부서":d,"총인원":n,"출근완료":comp,"지각":late,"결근":abst,"휴무":off,"미입력":ni,"출근율":r}
        stats.append(row)
        for k in tot:
            if k not in("부서","출근율"): tot[k]+=row.get(k,0)
    wt=tot["총인원"]-tot["휴무"]
    tot["출근율"]=f"{round(tot['출근완료']/wt*100,1)}%" if wt>0 else "0%"
    stats.append(tot)
    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric("출근율",tot["출근율"]); c2.metric("출근완료",f"{tot['출근완료']}명")
    c3.metric("지각",f"{tot['지각']}명"); c4.metric("결근",f"{tot['결근']}명")
    c5.metric("미입력",f"{tot['미입력']}명")
    st.markdown("---")
    st.dataframe(pd.DataFrame(stats),use_container_width=True,hide_index=True)
    st.markdown("---")
    detail=[]
    for e in emps:
        a=am.get(e["emp_no"],{})
        detail.append({"부서":e["dept_name"],"사번":e["emp_no"],"이름":e["name"],
                        "직급":e.get("position",""),"출근예정":e.get("scheduled_time",""),
                        "실제출근":a.get("actual_time",""),"상태":a.get("status","미입력")})
    ddf=pd.DataFrame(detail)
    st.dataframe(ddf,use_container_width=True,hide_index=True,height=400)
    st.download_button("📥 엑셀 다운로드",to_excel(ddf,"출근현황"),
                       f"출근현황_{ts}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════
#  📋 출근입력
# ═══════════════════════════════════════════════════════════════
def page_attendance():
    st.title("📋 부서별 출근 현황 입력")
    td=st.date_input("날짜",value=date.today()); ts=td.strftime("%Y-%m-%d")
    depts=sb_select("departments",order="sort_order,name")
    if not depts: st.warning("부서 없음"); return
    sel=st.selectbox("🏢 부서 선택",[d["name"] for d in depts])
    emps=sb_select("employees",f"dept_name=eq.{sel}&employment_status=eq.재직",order="emp_no")
    if not emps: st.info(f"{sel}에 직원 없음"); return
    atts=sb_select("attendance",f"date=eq.{ts}")
    am={a["emp_no"]:a for a in atts}
    st.markdown(f"**{sel}** — {len(emps)}명"); st.divider()
    nd={}
    for emp in emps:
        ex=am.get(emp["emp_no"],{})
        c1,c2,c3,c4=st.columns([2,1.5,1.2,1.5])
        c1.write(f"👤 {emp['name']}"); c2.write(emp["emp_no"])
        with c3: at=st.text_input("시간",value=ex.get("actual_time",""),placeholder="HH:MM",
                                   key=f"a_{emp['emp_no']}",label_visibility="collapsed")
        with c4:
            cs=ex.get("status","")
            idx=(STATUS_OPTIONS.index(cs)+1) if cs in STATUS_OPTIONS else 0
            s=st.selectbox("상태",[""]+STATUS_OPTIONS,index=idx,
                           key=f"s_{emp['emp_no']}",label_visibility="collapsed")
        nd[emp["emp_no"]]={"actual_time":at,"status":s}
    st.markdown("---")
    _,cb,_=st.columns([1,1,1])
    with cb:
        if st.button("💾 저장",use_container_width=True,type="primary"):
            # 시간 포맷 자동변환 적용
            recs=[]
            bad_times=[]
            for k,v in nd.items():
                if not v["status"]: continue
                t=fmt_time(v["actual_time"])
                if v["actual_time"] and t==v["actual_time"] and ":" not in t:
                    bad_times.append(k)  # 변환 실패
                recs.append({"date":ts,"emp_no":k,"actual_time":t,"status":v["status"]})
            if bad_times:
                st.error(f"시간 형식 오류: {', '.join(bad_times)} — 4자리 숫자(예:0900) 또는 HH:MM 형식으로 입력하세요.")
                return
            if recs:
                ok,msg=sb_upsert("attendance",recs,"date,emp_no")
                if ok: st.success(f"✅ {len(recs)}명 저장 완료!"); st.rerun()
                else: st.error(f"실패: {msg}")
            else: st.warning("상태를 선택하세요.")

# ═══════════════════════════════════════════════════════════════
#  🔍 이력조회
# ═══════════════════════════════════════════════════════════════
def page_history():
    st.title("🔍 출근 이력 조회")
    c1,c2,c3=st.columns(3)
    with c1: df_=st.date_input("시작일",value=date.today().replace(day=1))
    with c2: dt_=st.date_input("종료일",value=date.today())
    with c3:
        dpts=sb_select("departments",order="sort_order,name")
        sd=st.selectbox("부서",["전체"]+[d["name"] for d in dpts])
    logs=sb_select("attendance",f"date=gte.{df_}&date=lte.{dt_}",order="date.desc,emp_no")
    if not logs: st.info("이력 없음"); return
    emap={e["emp_no"]:e for e in sb_select("employees")}
    rows=[]
    for l in logs:
        e=emap.get(l["emp_no"],{})
        d=e.get("dept_name","-")
        if sd!="전체" and d!=sd: continue
        rows.append({"날짜":l["date"],"부서":d,"사번":l["emp_no"],"이름":e.get("name","-"),
                      "실제출근":l.get("actual_time",""),"상태":l.get("status","")})
    if rows:
        rdf=pd.DataFrame(rows)
        st.dataframe(rdf,use_container_width=True,hide_index=True,height=500)
        st.caption(f"총 {len(rows)}건")
        st.download_button("📥 엑셀",to_excel(rdf,"이력"),f"출근이력_{df_}_{dt_}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════════════════════
#  🏢 부서관리
# ═══════════════════════════════════════════════════════════════
def page_departments():
    st.title("🏢 부서 관리")
    t1,t2,t3=st.tabs(["📋 목록/수정","➕ 등록","🔀 순서 정렬"])
    with t1:
        all_depts=sb_select("departments",order="sort_order,name")
        for d in all_depts:
            ec=len(sb_select("employees",f"dept_name=eq.{d['name']}"))
            so=d.get('sort_order',0) or 0
            with st.expander(f"[{so}] 🏢 {d['name']} — {ec}명"):
                with st.form(f"dedit_{d['id']}"):
                    new_name=st.text_input("부서명",value=d['name'],key=f"dn_{d['id']}")
                    c1,c2=st.columns(2)
                    with c1:
                        if st.form_submit_button("✏️ 수정",type="primary"):
                            nn=new_name.strip()
                            if nn and nn!=d['name']:
                                ok,_=sb_update("departments",{"name":nn},f"id=eq.{d['id']}")
                                if ok:
                                    sb_update("employees",{"dept_name":nn},f"dept_name=eq.{d['name']}")
                                    st.success(f"'{d['name']}' → '{nn}' 변경 완료"); st.rerun()
                                else: st.error("중복 또는 오류")
                # 삭제: 소속 직원 유무 확인
                if ec==0:
                    del_key=f"del_confirm_{d['id']}"
                    if del_key not in st.session_state: st.session_state[del_key]=False
                    if not st.session_state[del_key]:
                        if st.button(f"🗑️ {d['name']} 삭제",key=f"dd_{d['id']}"):
                            st.session_state[del_key]=True; st.rerun()
                    else:
                        st.warning(f"정말 '{d['name']}' 부서를 삭제하시겠습니까?")
                        c1,c2=st.columns(2)
                        with c1:
                            if st.button("✅ 예, 삭제",key=f"dy_{d['id']}",type="primary"):
                                ok=sb_delete("departments",f"id=eq.{d['id']}")
                                if ok: st.success(f"'{d['name']}' 삭제 완료")
                                else: st.error("삭제 실패 — DB 제약조건 확인")
                                st.session_state[del_key]=False; st.rerun()
                        with c2:
                            if st.button("❌ 취소",key=f"dc_{d['id']}"):
                                st.session_state[del_key]=False; st.rerun()
                else:
                    st.caption(f"⚠️ 소속 직원 {ec}명 — 직원 이동/삭제 후 부서 삭제 가능")
    with t2:
        with st.form("nd"):
            nn=st.text_input("새 부서명")
            ns=st.number_input("표시 순서",value=len(all_depts)+1,min_value=1)
            if st.form_submit_button("➕ 등록",type="primary"):
                if nn:
                    ok,_=sb_insert("departments",{"name":nn.strip(),"sort_order":ns})
                    if ok: st.success("등록 완료"); st.rerun()
                    else: st.error("중복 또는 오류")
    with t3:
        st.subheader("🔀 부서 표시 순서")
        st.caption("숫자가 작을수록 위에 표시됩니다. 변경 후 '순서 저장'을 눌러주세요.")
        all_depts=sb_select("departments",order="sort_order,name")
        if all_depts:
            with st.form("dept_sort"):
                orders={}
                for d in all_depts:
                    c1,c2=st.columns([3,1])
                    c1.write(f"🏢 {d['name']}")
                    orders[d['id']]=c2.number_input("순번",value=d.get('sort_order',0) or 0,
                                                     min_value=0,key=f"so_{d['id']}",
                                                     label_visibility="collapsed")
                if st.form_submit_button("💾 순서 저장",type="primary"):
                    for did,so in orders.items():
                        sb_update("departments",{"sort_order":so},f"id=eq.{did}")
                    st.success("부서 순서가 저장되었습니다!"); st.rerun()

# ═══════════════════════════════════════════════════════════════
#  👤 직원관리
# ═══════════════════════════════════════════════════════════════
def page_employees():
    st.title("👤 직원 관리")
    t1,t2,t3,t4=st.tabs(["📋 직원 목록/수정","➕ 직원 추가","📥 엑셀 업로드","🗑️ 삭제"])

    with t1:
        depts=sb_select("departments",order="sort_order,name")
        df_=st.selectbox("부서",["전체"]+[d["name"] for d in depts],key="elf")
        if df_=="전체": emps=sb_select("employees",order="dept_name,emp_no")
        else: emps=sb_select("employees",f"dept_name=eq.{df_}",order="emp_no")
        if not emps: st.info("직원 없음"); return

        cols_show=["emp_no","name","dept_name","position","phone","hire_date",
                    "employment_status","health_cert_expiry","safety_training_date"]
        cols_label={"emp_no":"사번","name":"이름","dept_name":"부서","position":"직급",
                     "phone":"연락처","hire_date":"입사일","employment_status":"재직상태",
                     "health_cert_expiry":"보건증만료","safety_training_date":"안전교육일"}
        df=pd.DataFrame(emps)
        # 존재하는 컬럼만 표시
        avail=[c for c in cols_show if c in df.columns]
        show_df=df[avail].rename(columns=cols_label)
        st.dataframe(show_df,use_container_width=True,hide_index=True)
        st.caption(f"총 {len(emps)}명")

        # 직원 정보 수정
        st.markdown("---")
        st.subheader("✏️ 직원 정보 수정")
        emp_options=[f"{e['emp_no']} — {e['name']}" for e in emps]
        sel_emp_str=st.selectbox("수정할 직원 선택",emp_options,key="edit_sel")
        sel_no=sel_emp_str.split(" — ")[0] if sel_emp_str else ""
        sel_emp=next((e for e in emps if e["emp_no"]==sel_no),None)
        if sel_emp:
            pos_opts=["","사원","주임","대리","과장","차장","부장","이사","반장"]
            with st.form("edit_emp"):
                c1,c2,c3=st.columns(3)
                new_name=c1.text_input("이름",value=sel_emp.get("name",""))
                cur_pos=sel_emp.get("position","")
                pos_idx=pos_opts.index(cur_pos) if cur_pos in pos_opts else 0
                new_pos=c2.selectbox("직급",pos_opts,index=pos_idx)
                new_phone=c3.text_input("연락처",value=sel_emp.get("phone",""))
                c1,c2,c3=st.columns(3)
                new_hire=c1.text_input("입사일",value=sel_emp.get("hire_date",""))
                new_hce=c2.text_input("보건증만료일",value=sel_emp.get("health_cert_expiry",""))
                new_std=c3.text_input("안전교육이수일",value=sel_emp.get("safety_training_date",""))
                c1,c2=st.columns(2)
                new_sched=c1.text_input("출근예정시간",value=sel_emp.get("scheduled_time","09:00"))
                new_memo=c2.text_input("메모",value=sel_emp.get("memo",""))
                if st.form_submit_button("💾 수정 저장",type="primary"):
                    upd={"name":new_name.strip(),"position":new_pos,"phone":new_phone.strip(),
                         "hire_date":new_hire.strip(),"health_cert_expiry":new_hce.strip(),
                         "safety_training_date":new_std.strip(),"scheduled_time":new_sched.strip(),
                         "memo":new_memo.strip()}
                    ok,_=sb_update("employees",upd,f"emp_no=eq.{sel_no}")
                    if ok: st.success(f"{sel_no} 정보 수정 완료"); st.rerun()
                    else: st.error("수정 실패")

    with t2:
        st.subheader("➕ 직원 1명 추가")
        with st.form("add1"):
            c1,c2,c3=st.columns(3)
            ne=c1.text_input("사번 *")
            nn=c2.text_input("이름 *")
            nd=c3.selectbox("부서",[d["name"] for d in depts])
            c1,c2,c3=st.columns(3)
            np=c1.selectbox("직급",["","사원","주임","대리","과장","차장","부장","이사","반장"])
            nt=c2.text_input("출근예정시간",value="09:00")
            nh=c3.text_input("입사일",placeholder="2025-01-01")
            c1,c2=st.columns(2)
            nph=c1.text_input("연락처",placeholder="010-0000-0000")
            nct=c2.selectbox("계약유형",["정규직","계약직","일용직","파트타임"])
            if st.form_submit_button("➕ 등록",type="primary"):
                if ne and nn:
                    rec={"emp_no":ne.strip(),"name":nn.strip(),"dept_name":nd,
                         "position":np,"scheduled_time":nt,"hire_date":nh,
                         "phone":nph,"contract_type":nct,"employment_status":"재직"}
                    ok,msg=sb_insert("employees",[rec])
                    if ok: st.success("등록 완료"); st.rerun()
                    else: st.error(f"실패(사번 중복?): {msg}")

    with t3:
        st.subheader("📥 엑셀/CSV 일괄 업로드")
        st.markdown("컬럼: **사번, 이름, 부서, 출근예정시간** (필수) + 직급, 연락처 등 (선택)")
        up=st.file_uploader("파일 선택",type=["xlsx","csv"])
        if up:
            try:
                udf=pd.read_csv(up) if up.name.endswith('.csv') else pd.read_excel(up)
                req={"사번","이름","부서","출근예정시간"}
                if not req.issubset(set(udf.columns)):
                    st.error(f"필수 컬럼 누락: {req-set(udf.columns)}"); return
                st.dataframe(udf.head(10),use_container_width=True,hide_index=True)
                st.caption(f"총 {len(udf)}명")
                if st.button("🚀 DB 저장",type="primary"):
                    recs=[]
                    col_map={"사번":"emp_no","이름":"name","부서":"dept_name",
                             "출근예정시간":"scheduled_time","직급":"position",
                             "연락처":"phone","입사일":"hire_date","계약유형":"contract_type"}
                    for _,r in udf.iterrows():
                        rec={}
                        for kr,en in col_map.items():
                            if kr in udf.columns: rec[en]=str(r[kr]).strip()
                        rec["employment_status"]="재직"
                        recs.append(rec)
                    ok,msg=sb_upsert("employees",recs,"emp_no")
                    if ok: st.success(f"🎉 {len(recs)}명 저장!"); st.rerun()
                    else: st.error(f"실패: {msg}")
            except Exception as e: st.error(f"에러: {e}")

    with t4:
        st.subheader("🗑️ 직원 삭제")
        de=st.text_input("삭제할 사번",placeholder="ACC-01")
        if st.button("삭제 실행"):
            if de:
                ok=sb_delete("employees",f"emp_no=eq.{de.strip()}")
                if ok: st.success("삭제 완료"); st.rerun()
                else: st.error("사번 확인")

# ═══════════════════════════════════════════════════════════════
#  🦺 TBM 안전관리
# ═══════════════════════════════════════════════════════════════
def page_tbm():
    st.title("🦺 TBM 안전관리")
    t1,t2,t3=st.tabs(["📝 오늘의 TBM 작성","✅ 직원 확인 현황","📋 TBM 이력"])

    with t1:
        st.subheader("오늘의 TBM 작성")
        with st.form("tbm_write"):
            td=date.today().strftime("%Y-%m-%d")
            st.write(f"📅 날짜: **{td}**")
            dept=st.selectbox("대상 부서",[d["name"] for d in sb_select("departments",order="sort_order,name")])
            title=st.text_input("TBM 제목",placeholder="예: 수조 주변 미끄럼 주의")
            content=st.text_area("교육 내용",height=200,
                                 placeholder="1. 오늘의 작업 내용\n2. 위험 요인\n3. 안전 대책\n4. 주의사항")
            photo=st.file_uploader("📸 교육 현장 사진 (선택)",type=["jpg","jpeg","png"])
            if st.form_submit_button("💾 TBM 저장",type="primary"):
                if not title: st.error("제목을 입력하세요"); return
                photo_b64=""
                if photo:
                    raw=photo.read()
                    if len(raw)>3*1024*1024: st.error("사진 3MB 이하만 가능"); return
                    photo_b64=base64.b64encode(raw).decode()
                rec={"tbm_date":td,"department":dept,"title":title.strip(),
                     "content":content,"photo_data":photo_b64,"created_by":"admin"}
                ok,msg=sb_insert("tbm_records",[rec])
                if ok: st.success("✅ TBM이 등록되었습니다. 직원들이 확인할 수 있습니다!"); st.rerun()
                else: st.error(f"실패: {msg}")

    with t2:
        st.subheader("직원 확인 현황")
        today_tbms=sb_select("tbm_records",f"tbm_date=eq.{date.today().strftime('%Y-%m-%d')}",
                              order="created_at.desc")
        if not today_tbms:
            st.info("오늘 등록된 TBM이 없습니다. 먼저 TBM을 작성하세요.")
            return

        for tbm in today_tbms:
            with st.expander(f"📌 [{tbm['department']}] {tbm['title']}",expanded=True):
                st.markdown(tbm.get("content","").replace("\n","  \n"))
                if tbm.get("photo_data"):
                    try:
                        st.image(base64.b64decode(tbm["photo_data"]),caption="교육 현장",width=400)
                    except: pass

                # 확인한 직원 목록
                confs=sb_select("tbm_confirmations",f"tbm_id=eq.{tbm['id']}")
                conf_nos={c["emp_no"] for c in confs}

                # 해당 부서 직원
                dept_emps=sb_select("employees",
                                    f"dept_name=eq.{tbm['department']}&employment_status=eq.재직",
                                    order="emp_no")

                st.markdown(f"**확인 현황: {len(conf_nos)}/{len(dept_emps)}명**")
                for emp in dept_emps:
                    c1,c2=st.columns([3,1])
                    if emp["emp_no"] in conf_nos:
                        c1.write(f"✅ {emp['name']} ({emp['emp_no']})")
                        c2.write("확인 완료")
                    else:
                        c1.write(f"⬜ {emp['name']} ({emp['emp_no']})")
                        with c2:
                            if st.button("확인",key=f"cf_{tbm['id']}_{emp['emp_no']}"):
                                ok,_=sb_insert("tbm_confirmations",[{
                                    "tbm_id":tbm["id"],"emp_no":emp["emp_no"],
                                    "emp_name":emp["name"]
                                }])
                                if ok: st.rerun()

                # 미확인자 알림
                uncf=[e for e in dept_emps if e["emp_no"] not in conf_nos]
                if uncf:
                    st.warning(f"⚠️ 미확인 {len(uncf)}명: {', '.join(e['name'] for e in uncf)}")

    with t3:
        st.subheader("TBM 이력")
        c1,c2=st.columns(2)
        with c1: hf=st.date_input("시작",value=date.today().replace(day=1),key="tbmf")
        with c2: ht=st.date_input("종료",value=date.today(),key="tbmt")
        hist=sb_select("tbm_records",f"tbm_date=gte.{hf}&tbm_date=lte.{ht}",
                        order="tbm_date.desc,created_at.desc")
        if hist:
            for h in hist:
                confs=sb_select("tbm_confirmations",f"tbm_id=eq.{h['id']}")
                dept_emps=sb_select("employees",
                                    f"dept_name=eq.{h['department']}&employment_status=eq.재직")
                rate=f"{len(confs)}/{len(dept_emps)}" if dept_emps else "0/0"
                with st.expander(f"📅 {h['tbm_date']} [{h['department']}] {h['title']} — 확인 {rate}"):
                    st.markdown(h.get("content","").replace("\n","  \n"))
                    if h.get("photo_data"):
                        try: st.image(base64.b64decode(h["photo_data"]),width=300)
                        except: pass
                    if confs:
                        st.markdown("**확인자:** " + ", ".join(
                            f"{c.get('emp_name','')}({c['emp_no']})" for c in confs))
        else:
            st.info("해당 기간 TBM 이력 없음")

# ═══════════════════════════════════════════════════════════════
#  💰 급여관리 / ⚙️ 계정관리
# ═══════════════════════════════════════════════════════════════
def page_salary():
    st.title("💰 급여 관리")
    st.info("급여 입력/저장 기능은 데이터 안정화 후 추가 예정입니다.")
    emps=sb_select("employees","employment_status=eq.재직",order="dept_name,emp_no")
    if emps:
        df=pd.DataFrame(emps)
        cols=["emp_no","name","dept_name","position","base_salary","allowance","pay_type","bank_name"]
        avail=[c for c in cols if c in df.columns]
        show=df[avail].rename(columns={"emp_no":"사번","name":"이름","dept_name":"부서",
                "position":"직급","base_salary":"기본급","allowance":"수당",
                "pay_type":"급여유형","bank_name":"은행"})
        st.dataframe(show,use_container_width=True,hide_index=True)

DEFAULT_MENU_ORDER=["📊 전사현황","📋 출근입력","🔍 이력조회","🏢 부서관리",
                    "👤 직원관리","🦺 TBM 안전관리","💰 급여관리","⚙️ 시스템관리"]
MENU_MAP={"📊 전사현황":page_dashboard,"📋 출근입력":page_attendance,
          "🔍 이력조회":page_history,"🏢 부서관리":page_departments,
          "👤 직원관리":page_employees,"🦺 TBM 안전관리":page_tbm,
          "💰 급여관리":page_salary,"⚙️ 시스템관리":None}

def get_menu_order():
    if 'menu_order' not in st.session_state:
        st.session_state['menu_order']=list(DEFAULT_MENU_ORDER)
    return st.session_state['menu_order']

def page_settings():
    st.title("⚙️ 시스템 관리")
    t1,t2=st.tabs(["📊 시스템 상태","🔀 메뉴 순서 관리"])
    with t1:
        if sb_health(): st.success("✅ Supabase 연결 정상")
        else: st.error("❌ 연결 실패")
        ec=len(sb_select("employees","employment_status=eq.재직"))
        dc=len(sb_select("departments"))
        ac=len(sb_select("attendance"))
        tc=len(sb_select("tbm_records"))
        c1,c2,c3,c4=st.columns(4)
        c1.metric("부서",dc); c2.metric("직원(재직)",ec)
        c3.metric("출근기록",ac); c4.metric("TBM기록",tc)
    with t2:
        st.subheader("🔀 메뉴 순서 관리")
        st.caption("각 메뉴의 순번을 지정하고 '적용' 버튼을 누르세요.")
        cur=get_menu_order()
        with st.form("menu_order_form"):
            new_orders={}
            for i,m in enumerate(cur):
                c1,c2=st.columns([3,1])
                c1.write(m)
                new_orders[m]=c2.number_input("순번",value=i+1,min_value=1,max_value=8,
                                              key=f"mo_{m}",label_visibility="collapsed")
            if st.form_submit_button("✅ 순서 적용",type="primary"):
                sorted_menus=sorted(new_orders.keys(),key=lambda x:new_orders[x])
                st.session_state['menu_order']=sorted_menus
                st.success("메뉴 순서가 변경되었습니다!"); st.rerun()
        if st.button("🔄 기본 순서로 초기화"):
            st.session_state['menu_order']=list(DEFAULT_MENU_ORDER)
            st.rerun()

# ═══════════════════════════════════════════════════════════════
#  메인
# ═══════════════════════════════════════════════════════════════
def main():
    if not st.session_state['logged_in']:
        if not sb_health(): st.error("❌ Supabase 연결 실패")
        show_login(); return
    init_departments()
    ordered=get_menu_order()
    st.sidebar.markdown("### 🐟 보물섬수산 HR")
    st.sidebar.markdown("---")
    menu=st.sidebar.radio("메뉴",ordered)
    st.sidebar.markdown("---")
    if st.sidebar.button("🔓 로그아웃",use_container_width=True):
        st.session_state['logged_in']=False; st.rerun()
    st.sidebar.caption("v2.2")
    page_fn=MENU_MAP.get(menu)
    if page_fn: page_fn()
    else: page_settings()

if __name__=="__main__": main()
