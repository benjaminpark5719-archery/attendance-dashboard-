"""
=================================================================
보물섬수산 출근현황 관리 대시보드 v2.0
- 관리자 인증 (총괄관리자 / 부서관리자)
- 부서·직원 CRUD
- 출근 기록 DB 저장
- 급여 관리
=================================================================
"""

import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from database import (
    init_db,
    # Auth
    authenticate_user, create_user, change_password,
    get_all_users, toggle_user_active, reset_user_password,
    # Department
    get_all_departments, create_department, update_department, delete_department,
    # Employee
    get_employees, create_employee, update_employee,
    deactivate_employee, reactivate_employee,
    # Attendance
    get_attendance, save_attendance, get_attendance_stats,
    get_attendance_history,
    # Salary
    get_salary_info, save_salary_info, get_salary_summary,
)

# ─── Page Config ─────────────────────────────────────────────────

st.set_page_config(
    page_title="보물섬수산 출근현황",
    page_icon="🐟",
    layout="wide",
    initial_sidebar_state="expanded"
)

STATUS_OPTIONS = ["출근완료", "지각", "결근", "휴무", "조퇴"]
POSITION_OPTIONS = ["", "사원", "주임", "대리", "과장", "차장", "부장", "이사", "상무", "전무", "대표"]
PAY_TYPES = ["월급", "일급", "시급"]

# ─── Custom CSS ──────────────────────────────────────────────────

st.markdown("""
<style>
    .main-header {
        font-size: 1.6rem; font-weight: 700;
        color: #1a365d; margin-bottom: 0.5rem;
    }
    .stat-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white; padding: 1.2rem; border-radius: 12px;
        text-align: center; margin-bottom: 1rem;
    }
    .stat-card h3 { margin: 0; font-size: 2rem; }
    .stat-card p { margin: 0.3rem 0 0; font-size: 0.85rem; opacity: 0.9; }
    .login-box {
        max-width: 400px; margin: 5rem auto; padding: 2rem;
        border: 1px solid #e2e8f0; border-radius: 16px;
        background: white;
    }
    div[data-testid="stSidebar"] { background: #f7fafc; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 20px; border-radius: 8px;
    }
</style>
""", unsafe_allow_html=True)


# ─── Session Helpers ─────────────────────────────────────────────

def is_logged_in():
    return st.session_state.get('user') is not None

def current_user():
    return st.session_state.get('user', {})

def is_super_admin():
    return current_user().get('role') == 'super_admin'

def logout():
    for key in ['user', 'page']:
        st.session_state.pop(key, None)
    st.rerun()


# ─── Excel Export ────────────────────────────────────────────────

def export_to_excel(df, sheet_name='데이터'):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        hfill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)
    output.seek(0)
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════
#  LOGIN PAGE
# ═══════════════════════════════════════════════════════════════

def page_login():
    st.markdown("<div style='text-align:center; margin-top:3rem;'>"
                "<h1>🐟 보물섬수산</h1>"
                "<h3 style='color:#64748b;'>출근현황 관리 대시보드</h3>"
                "</div>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("---")
        st.markdown("#### 🔐 관리자 로그인")

        with st.form("login_form"):
            username = st.text_input("아이디", placeholder="admin")
            password = st.text_input("비밀번호", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("로그인", use_container_width=True, type="primary")

        if submitted:
            if not username or not password:
                st.error("아이디와 비밀번호를 입력하세요.")
            else:
                user = authenticate_user(username, password)
                if user:
                    st.session_state['user'] = user
                    st.success(f"✅ {user['display_name']}님 환영합니다!")
                    st.rerun()
                else:
                    st.error("❌ 아이디 또는 비밀번호가 올바르지 않습니다.")

        st.caption("초기 관리자: admin / admin1234!")


# ═══════════════════════════════════════════════════════════════
#  PASSWORD CHANGE (강제 변경)
# ═══════════════════════════════════════════════════════════════

def page_change_password():
    st.markdown("### 🔑 비밀번호 변경")
    st.warning("보안을 위해 초기 비밀번호를 변경해주세요.")

    with st.form("pw_change"):
        new_pw = st.text_input("새 비밀번호", type="password",
                               help="8자 이상 권장")
        new_pw2 = st.text_input("새 비밀번호 확인", type="password")
        submitted = st.form_submit_button("변경하기", type="primary")

    if submitted:
        if len(new_pw) < 6:
            st.error("비밀번호는 6자 이상이어야 합니다.")
        elif new_pw != new_pw2:
            st.error("비밀번호가 일치하지 않습니다.")
        else:
            if change_password(current_user()['id'], new_pw):
                st.session_state['user']['must_change_pw'] = False
                st.success("✅ 비밀번호가 변경되었습니다!")
                st.rerun()


# ═══════════════════════════════════════════════════════════════
#  SIDEBAR NAVIGATION
# ═══════════════════════════════════════════════════════════════

def render_sidebar():
    user = current_user()
    st.sidebar.markdown(f"### 🐟 보물섬수산")
    st.sidebar.markdown(f"**{user['display_name']}** ({user['role'].replace('_', ' ').title()})")
    st.sidebar.markdown("---")

    if is_super_admin():
        menu_items = {
            "📊 전사 출근 현황": "dashboard",
            "📋 부서별 출근 입력": "dept_attendance",
            "📅 출근 이력 조회": "history",
            "🏢 부서 관리": "departments",
            "👥 직원 관리": "employees",
            "💰 급여 관리": "salary",
            "🔧 관리자 계정": "users",
        }
    else:
        menu_items = {
            "📋 출근 현황 입력": "dept_attendance",
            "📅 출근 이력 조회": "history",
        }

    for label, page_key in menu_items.items():
        if st.sidebar.button(label, use_container_width=True,
                             type="primary" if st.session_state.get('page') == page_key else "secondary"):
            st.session_state['page'] = page_key
            st.rerun()

    st.sidebar.markdown("---")
    if st.sidebar.button("🔓 로그아웃", use_container_width=True):
        logout()

    st.sidebar.markdown("---")
    st.sidebar.caption("v2.0 | 보물섬수산 HR시스템")


# ═══════════════════════════════════════════════════════════════
#  PAGE: 전사 출근 현황 대시보드
# ═══════════════════════════════════════════════════════════════

def page_dashboard():
    st.markdown("## 📊 전사 출근 현황 대시보드")
    target_date = st.date_input("📅 기준일", value=date.today())

    stats = get_attendance_stats(target_date)
    if not stats:
        st.info("등록된 부서 또는 직원이 없습니다.")
        return

    totals = stats[-1] if stats and stats[-1].get('부서') == '합계' else {}

    # KPI 카드
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("📊 출근율", totals.get('출근율', '0%'))
    c2.metric("✅ 출근완료", f"{totals.get('출근완료', 0)}명")
    c3.metric("⏰ 지각", f"{totals.get('지각', 0)}명")
    c4.metric("❌ 결근", f"{totals.get('결근', 0)}명")
    c5.metric("📝 미입력", f"{totals.get('미입력', 0)}명")

    st.markdown("---")

    # 부서별 상세
    st.markdown("### 🏢 부서별 상세 현황")
    stats_df = pd.DataFrame(stats)
    st.dataframe(stats_df, use_container_width=True, hide_index=True)

    # 전체 직원 상세 리스트
    st.markdown("---")
    st.markdown("### 👥 전체 직원 상세 리스트")

    depts = get_all_departments(active_only=True)
    dept_names = [d['name'] for d in depts]

    c1, c2 = st.columns(2)
    with c1:
        filter_dept = st.multiselect("부서 필터:", dept_names)
    with c2:
        filter_status = st.multiselect("상태 필터:", STATUS_OPTIONS + ["미입력"])

    records = get_attendance(target_date)
    if filter_dept:
        records = [r for r in records if r['department_name'] in filter_dept]
    if filter_status:
        records = [r for r in records
                   if r['status'] in filter_status or
                   (not r['status'] and '미입력' in filter_status)]

    if records:
        df = pd.DataFrame(records)
        display_cols = ['department_name', 'emp_number', 'name', 'position',
                        'scheduled_time', 'actual_time', 'status', 'note']
        rename = {'department_name': '부서', 'emp_number': '사번', 'name': '이름',
                  'position': '직급', 'scheduled_time': '출근예정',
                  'actual_time': '실제출근', 'status': '상태', 'note': '비고'}
        display_df = df[display_cols].rename(columns=rename)
        st.dataframe(display_df, use_container_width=True, hide_index=True, height=400)

        # 엑셀 다운로드
        st.download_button(
            "📥 엑셀 다운로드",
            export_to_excel(display_df, '출근현황'),
            f"출근현황_{target_date.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("해당 조건의 데이터가 없습니다.")


# ═══════════════════════════════════════════════════════════════
#  PAGE: 부서별 출근 입력
# ═══════════════════════════════════════════════════════════════

def page_dept_attendance():
    user = current_user()
    st.markdown("## 📋 출근 현황 입력")

    target_date = st.date_input("📅 날짜", value=date.today())

    depts = get_all_departments(active_only=True)
    if is_super_admin():
        dept_names = [d['name'] for d in depts]
        selected_dept_name = st.selectbox("🏢 부서 선택", dept_names)
        selected_dept = next(d for d in depts if d['name'] == selected_dept_name)
    else:
        # 부서 관리자: 본인 부서만
        selected_dept = next((d for d in depts if d['id'] == user.get('department_id')), None)
        if not selected_dept:
            st.error("소속 부서가 지정되지 않았습니다. 총괄 관리자에게 문의하세요.")
            return
        st.info(f"🏢 소속 부서: **{selected_dept['name']}**")

    records = get_attendance(target_date, department_id=selected_dept['id'])
    if not records:
        st.info("해당 부서에 등록된 직원이 없습니다.")
        return

    st.markdown(f"**{selected_dept['name']}** — 총 {len(records)}명")
    st.markdown("---")

    # 헤더
    cols = st.columns([1.5, 1, 1.2, 1.2, 1.5, 1.5])
    headers = ["이름", "직급", "출근예정", "실제출근", "근무상태", "비고"]
    for col, h in zip(cols, headers):
        col.markdown(f"**{h}**")
    st.divider()

    # 직원별 입력 행
    new_values = {}
    for rec in records:
        cols = st.columns([1.5, 1, 1.2, 1.2, 1.5, 1.5])
        with cols[0]:
            st.write(f"👤 {rec['name']}")
        with cols[1]:
            st.write(rec['position'] or '-')
        with cols[2]:
            st.write(rec['scheduled_time'])
        with cols[3]:
            actual = st.text_input(
                "시간", value=rec['actual_time'],
                placeholder="HH:MM",
                key=f"at_{rec['employee_id']}",
                label_visibility="collapsed"
            )
        with cols[4]:
            idx = STATUS_OPTIONS.index(rec['status']) if rec['status'] in STATUS_OPTIONS else 0
            status = st.selectbox(
                "상태", [""] + STATUS_OPTIONS, 
                index=(STATUS_OPTIONS.index(rec['status']) + 1) if rec['status'] in STATUS_OPTIONS else 0,
                key=f"st_{rec['employee_id']}",
                label_visibility="collapsed"
            )
        with cols[5]:
            note = st.text_input(
                "비고", value=rec['note'],
                key=f"nt_{rec['employee_id']}",
                label_visibility="collapsed"
            )
        new_values[rec['employee_id']] = {
            'employee_id': rec['employee_id'],
            'department_id': rec['department_id'],
            'scheduled_time': rec['scheduled_time'],
            'actual_time': actual,
            'status': status,
            'note': note
        }

    st.markdown("---")
    c1, c2, c3 = st.columns([1, 1, 1])
    with c2:
        if st.button("💾 저장", use_container_width=True, type="primary"):
            save_records = list(new_values.values())
            save_attendance(target_date, save_records, registered_by_id=user.get('id'))
            st.success(f"✅ {target_date} {selected_dept['name']} 출근 데이터가 저장되었습니다!")
            st.rerun()


# ═══════════════════════════════════════════════════════════════
#  PAGE: 출근 이력 조회
# ═══════════════════════════════════════════════════════════════

def page_history():
    st.markdown("## 📅 출근 이력 조회")

    c1, c2, c3 = st.columns(3)
    with c1:
        date_from = st.date_input("시작일", value=date.today() - timedelta(days=30))
    with c2:
        date_to = st.date_input("종료일", value=date.today())
    with c3:
        depts = get_all_departments(active_only=True)
        user = current_user()
        if is_super_admin():
            dept_options = ["전체"] + [d['name'] for d in depts]
            sel = st.selectbox("부서", dept_options)
            dept_id = next((d['id'] for d in depts if d['name'] == sel), None)
        else:
            dept_id = user.get('department_id')
            dept_name = next((d['name'] for d in depts if d['id'] == dept_id), '소속부서')
            st.text_input("부서", value=dept_name, disabled=True)

    logs = get_attendance_history(
        department_id=dept_id,
        date_from=date_from, date_to=date_to
    )

    if logs:
        df = pd.DataFrame(logs)
        rename = {
            'log_date': '날짜', 'emp_number': '사번', 'name': '이름',
            'department': '부서', 'position': '직급',
            'scheduled_time': '출근예정', 'actual_time': '실제출근',
            'status': '상태', 'note': '비고', 'registered_by': '등록자'
        }
        display_df = df.rename(columns=rename)
        cols_order = ['날짜', '부서', '사번', '이름', '직급', '출근예정', '실제출근', '상태', '비고', '등록자']
        display_df = display_df[[c for c in cols_order if c in display_df.columns]]

        st.dataframe(display_df, use_container_width=True, hide_index=True, height=500)
        st.caption(f"총 {len(logs)}건")

        st.download_button(
            "📥 엑셀 다운로드",
            export_to_excel(display_df, '출근이력'),
            f"출근이력_{date_from}_{date_to}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("해당 기간의 출근 이력이 없습니다.")


# ═══════════════════════════════════════════════════════════════
#  PAGE: 부서 관리 (CRUD)
# ═══════════════════════════════════════════════════════════════

def page_departments():
    st.markdown("## 🏢 부서 관리")

    tab1, tab2 = st.tabs(["📋 부서 목록", "➕ 부서 등록"])

    with tab1:
        depts = get_all_departments()
        if depts:
            for dept in depts:
                with st.expander(
                    f"{'✅' if dept['is_active'] else '⛔'} {dept['name']} "
                    f"({dept['code']}) — 직원 {dept['employee_count']}명",
                    expanded=False
                ):
                    with st.form(f"dept_edit_{dept['id']}"):
                        c1, c2, c3 = st.columns(3)
                        new_name = c1.text_input("부서명", value=dept['name'])
                        new_code = c2.text_input("코드", value=dept['code'])
                        new_order = c3.number_input("정렬순서", value=dept['sort_order'])

                        c1, c2, c3 = st.columns(3)
                        with c1:
                            if st.form_submit_button("💾 수정", type="primary"):
                                ok, msg = update_department(dept['id'], new_name, new_code, new_order)
                                if ok:
                                    st.success(msg)
                                    st.rerun()
                                else:
                                    st.error(msg)

                    if dept['is_active']:
                        if st.button(f"🗑️ 비활성화", key=f"del_dept_{dept['id']}"):
                            ok, msg = delete_department(dept['id'])
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.error(msg)
        else:
            st.info("등록된 부서가 없습니다.")

    with tab2:
        with st.form("new_dept"):
            c1, c2, c3 = st.columns(3)
            name = c1.text_input("부서명 *")
            code = c2.text_input("부서코드 *", help="영문 3자 권장 (예: ACC)")
            order = c3.number_input("정렬순서", value=0)

            if st.form_submit_button("➕ 등록", type="primary"):
                if not name or not code:
                    st.error("부서명과 코드는 필수입니다.")
                else:
                    ok, msg = create_department(name, code.upper(), order)
                    if ok:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)


# ═══════════════════════════════════════════════════════════════
#  PAGE: 직원 관리 (CRUD)
# ═══════════════════════════════════════════════════════════════

def page_employees():
    st.markdown("## 👥 직원 관리")

    tab1, tab2, tab3 = st.tabs(["📋 직원 목록", "➕ 직원 등록", "🔄 퇴사자 관리"])

    depts = get_all_departments(active_only=True)
    dept_map = {d['id']: d['name'] for d in depts}

    with tab1:
        c1, c2 = st.columns([1, 3])
        with c1:
            filter_dept_name = st.selectbox("부서 필터", ["전체"] + [d['name'] for d in depts], key="emp_filter")

        dept_id = next((d['id'] for d in depts if d['name'] == filter_dept_name), None)
        employees = get_employees(department_id=dept_id, active_only=True)

        if employees:
            st.caption(f"총 {len(employees)}명")
            for emp in employees:
                with st.expander(
                    f"👤 {emp['name']} ({emp['emp_number']}) — "
                    f"{emp['department_name']} / {emp['position'] or '-'}",
                    expanded=False
                ):
                    with st.form(f"emp_edit_{emp['id']}"):
                        c1, c2, c3 = st.columns(3)
                        new_num = c1.text_input("사번", value=emp['emp_number'])
                        new_name = c2.text_input("이름", value=emp['name'])
                        new_pos = c3.selectbox(
                            "직급", POSITION_OPTIONS,
                            index=POSITION_OPTIONS.index(emp['position']) if emp['position'] in POSITION_OPTIONS else 0
                        )

                        c1, c2, c3 = st.columns(3)
                        dept_names = [d['name'] for d in depts]
                        new_dept = c1.selectbox(
                            "소속부서", dept_names,
                            index=dept_names.index(emp['department_name']) if emp['department_name'] in dept_names else 0,
                            key=f"edept_{emp['id']}"
                        )
                        new_phone = c2.text_input("연락처", value=emp['phone'] or '')
                        new_time = c3.text_input("출근예정시간", value=emp['scheduled_time'])

                        c1, c2 = st.columns(2)
                        new_hire = c1.date_input("입사일",
                                                 value=emp['hire_date'] if emp['hire_date'] else date.today(),
                                                 key=f"hire_{emp['id']}")

                        c1, c2, c3 = st.columns(3)
                        with c1:
                            if st.form_submit_button("💾 수정", type="primary"):
                                new_dept_id = next(d['id'] for d in depts if d['name'] == new_dept)
                                ok, msg = update_employee(
                                    emp['id'],
                                    emp_number=new_num, name=new_name,
                                    position=new_pos,
                                    department_id=new_dept_id,
                                    phone=new_phone,
                                    scheduled_time=new_time,
                                    hire_date=new_hire
                                )
                                if ok:
                                    st.success(msg)
                                    st.rerun()
                                else:
                                    st.error(msg)

                    if st.button(f"🚪 퇴사 처리", key=f"resign_{emp['id']}"):
                        ok, msg = deactivate_employee(emp['id'])
                        if ok:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
        else:
            st.info("등록된 직원이 없습니다.")

    with tab2:
        st.markdown("### ➕ 신규 직원 등록")
        with st.form("new_emp"):
            c1, c2, c3 = st.columns(3)
            emp_num = c1.text_input("사원번호 *", placeholder="예: EMP001")
            emp_name = c2.text_input("이름 *")
            emp_pos = c3.selectbox("직급", POSITION_OPTIONS)

            c1, c2, c3 = st.columns(3)
            emp_dept = c1.selectbox("소속부서 *", [d['name'] for d in depts])
            emp_phone = c2.text_input("연락처", placeholder="010-0000-0000")
            emp_time = c3.text_input("출근예정시간", value="09:00")

            emp_hire = st.date_input("입사일", value=date.today())

            if st.form_submit_button("➕ 등록", type="primary"):
                if not emp_num or not emp_name:
                    st.error("사원번호와 이름은 필수입니다.")
                else:
                    dept_id = next(d['id'] for d in depts if d['name'] == emp_dept)
                    ok, msg = create_employee(
                        emp_number=emp_num, name=emp_name,
                        department_id=dept_id, position=emp_pos,
                        phone=emp_phone, scheduled_time=emp_time,
                        hire_date=emp_hire
                    )
                    if ok:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)

    with tab3:
        st.markdown("### 🔄 퇴사자 목록")
        resigned = get_employees(active_only=False)
        resigned = [e for e in resigned if not e['is_active']]
        if resigned:
            for emp in resigned:
                c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
                c1.write(f"👤 {emp['name']} ({emp['emp_number']})")
                c2.write(emp['department_name'])
                c3.write(f"퇴사일: {emp['resign_date'] or '-'}")
                with c4:
                    if st.button("♻️ 복직", key=f"react_{emp['id']}"):
                        ok, msg = reactivate_employee(emp['id'])
                        if ok:
                            st.success(msg)
                            st.rerun()
        else:
            st.info("퇴사자가 없습니다.")


# ═══════════════════════════════════════════════════════════════
#  PAGE: 급여 관리
# ═══════════════════════════════════════════════════════════════

def page_salary():
    st.markdown("## 💰 급여 관리")

    tab1, tab2 = st.tabs(["📋 급여 현황", "✏️ 급여 입력/수정"])

    depts = get_all_departments(active_only=True)

    with tab1:
        dept_filter = st.selectbox("부서", ["전체"] + [d['name'] for d in depts], key="sal_dept")
        dept_id = next((d['id'] for d in depts if d['name'] == dept_filter), None)

        summary = get_salary_summary(department_id=dept_id)
        if summary:
            df = pd.DataFrame(summary)
            display_df = df[['emp_number', 'name', 'department', 'position',
                             'base_salary', 'total_allowance', 'total_deduction',
                             'net_salary', 'pay_type']].rename(columns={
                'emp_number': '사번', 'name': '이름', 'department': '부서',
                'position': '직급', 'base_salary': '기본급',
                'total_allowance': '총수당', 'total_deduction': '총공제',
                'net_salary': '실수령액', 'pay_type': '급여유형'
            })

            # 금액 컬럼 포맷
            for col in ['기본급', '총수당', '총공제', '실수령액']:
                display_df[col] = display_df[col].apply(lambda x: f"{x:,}" if x else "0")

            st.dataframe(display_df, use_container_width=True, hide_index=True)

            # 합계
            total_base = sum(s['base_salary'] for s in summary)
            total_allow = sum(s['total_allowance'] for s in summary)
            total_deduct = sum(s['total_deduction'] for s in summary)
            total_net = sum(s['net_salary'] for s in summary)

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("총 기본급", f"{total_base:,}원")
            c2.metric("총 수당", f"{total_allow:,}원")
            c3.metric("총 공제", f"{total_deduct:,}원")
            c4.metric("총 실수령액", f"{total_net:,}원")

            st.download_button(
                "📥 급여현황 엑셀 다운로드",
                export_to_excel(display_df, '급여현황'),
                f"급여현황_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("급여 데이터가 없습니다.")

    with tab2:
        st.markdown("### ✏️ 개인별 급여 입력/수정")

        # 직원 선택
        dept_sel = st.selectbox("부서 선택", [d['name'] for d in depts], key="sal_dept_sel")
        dept_id = next(d['id'] for d in depts if d['name'] == dept_sel)
        employees = get_employees(department_id=dept_id)

        if not employees:
            st.info("해당 부서에 직원이 없습니다.")
            return

        emp_names = [f"{e['name']} ({e['emp_number']})" for e in employees]
        sel_emp_name = st.selectbox("직원 선택", emp_names, key="sal_emp_sel")
        sel_emp = employees[emp_names.index(sel_emp_name)]

        # 기존 급여 정보 로드
        existing = get_salary_info(sel_emp['id'])

        with st.form("salary_form"):
            st.markdown("#### 💵 지급 항목")
            c1, c2, c3 = st.columns(3)
            base = c1.number_input("기본급", value=existing['base_salary'] if existing else 0,
                                   step=10000, min_value=0)
            pos_allow = c2.number_input("직책수당", value=existing['position_allowance'] if existing else 0,
                                        step=10000, min_value=0)
            meal = c3.number_input("식대", value=existing['meal_allowance'] if existing else 200000,
                                   step=10000, min_value=0)

            c1, c2, c3 = st.columns(3)
            transport = c1.number_input("교통비", value=existing['transport_allowance'] if existing else 0,
                                        step=10000, min_value=0)
            overtime = c2.number_input("연장근무수당", value=existing['overtime_allowance'] if existing else 0,
                                       step=10000, min_value=0)
            other_a = c3.number_input("기타수당", value=existing['other_allowance'] if existing else 0,
                                      step=10000, min_value=0)

            st.markdown("---")
            st.markdown("#### 📤 공제 항목")
            c1, c2, c3 = st.columns(3)
            pension = c1.number_input("국민연금", value=existing['national_pension'] if existing else 0,
                                      step=1000, min_value=0)
            health = c2.number_input("건강보험", value=existing['health_insurance'] if existing else 0,
                                     step=1000, min_value=0)
            employ = c3.number_input("고용보험", value=existing['employment_insurance'] if existing else 0,
                                     step=1000, min_value=0)

            c1, c2, c3 = st.columns(3)
            income_t = c1.number_input("소득세", value=existing['income_tax'] if existing else 0,
                                       step=1000, min_value=0)
            local_t = c2.number_input("지방소득세", value=existing['local_tax'] if existing else 0,
                                      step=1000, min_value=0)
            other_d = c3.number_input("기타공제", value=existing['other_deduction'] if existing else 0,
                                      step=1000, min_value=0)

            st.markdown("---")
            c1, c2, c3 = st.columns(3)
            pay_type = c1.selectbox("급여유형", PAY_TYPES,
                                    index=PAY_TYPES.index(existing['pay_type']) if existing and existing['pay_type'] in PAY_TYPES else 0)
            eff_date = c2.date_input("적용일",
                                     value=existing['effective_date'] if existing and existing['effective_date'] else date.today())
            notes = c3.text_input("메모", value=existing['notes'] if existing else '')

            # 실수령액 미리 계산
            total_allow = pos_allow + meal + transport + overtime + other_a
            total_deduct = pension + health + employ + income_t + local_t + other_d
            net = base + total_allow - total_deduct

            st.markdown(f"### 💰 실수령액: **{net:,}원**")
            st.caption(f"기본급 {base:,} + 수당 {total_allow:,} - 공제 {total_deduct:,}")

            if st.form_submit_button("💾 저장", type="primary"):
                save_salary_info(sel_emp['id'], {
                    'base_salary': base,
                    'position_allowance': pos_allow,
                    'meal_allowance': meal,
                    'transport_allowance': transport,
                    'overtime_allowance': overtime,
                    'other_allowance': other_a,
                    'national_pension': pension,
                    'health_insurance': health,
                    'employment_insurance': employ,
                    'income_tax': income_t,
                    'local_tax': local_t,
                    'other_deduction': other_d,
                    'pay_type': pay_type,
                    'effective_date': eff_date,
                    'notes': notes,
                })
                st.success(f"✅ {sel_emp['name']}님의 급여 정보가 저장되었습니다.")
                st.rerun()


# ═══════════════════════════════════════════════════════════════
#  PAGE: 관리자 계정 관리
# ═══════════════════════════════════════════════════════════════

def page_users():
    st.markdown("## 🔧 관리자 계정 관리")

    tab1, tab2 = st.tabs(["📋 계정 목록", "➕ 계정 생성"])

    with tab1:
        users = get_all_users()
        if users:
            for u in users:
                icon = "🟢" if u['is_active'] else "🔴"
                role_label = "총괄관리자" if u['role'] == 'super_admin' else "부서관리자"
                with st.expander(
                    f"{icon} {u['display_name']} (@{u['username']}) — {role_label} "
                    f"{'/ ' + u['department_name'] if u['department_name'] != '-' else ''}",
                    expanded=False
                ):
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        new_status = not u['is_active']
                        label = "✅ 활성화" if new_status else "⛔ 비활성화"
                        if st.button(label, key=f"toggle_{u['id']}"):
                            toggle_user_active(u['id'], new_status)
                            st.rerun()
                    with c2:
                        if st.button("🔑 비밀번호 초기화", key=f"reset_{u['id']}"):
                            reset_user_password(u['id'], "reset1234!")
                            st.success(f"비밀번호가 'reset1234!'로 초기화되었습니다.")

    with tab2:
        st.markdown("### ➕ 새 관리자 계정 생성")
        depts = get_all_departments(active_only=True)

        with st.form("new_user"):
            c1, c2 = st.columns(2)
            username = c1.text_input("아이디 *", placeholder="예: manager_acc")
            display_name = c2.text_input("표시명 *", placeholder="예: 회계부 관리자")

            c1, c2 = st.columns(2)
            password = c1.text_input("초기 비밀번호 *", type="password", value="admin1234!")
            role = c2.selectbox("권한", ["dept_admin", "super_admin"],
                                format_func=lambda x: "부서관리자" if x == "dept_admin" else "총괄관리자")

            dept_name = st.selectbox("소속부서 (부서관리자인 경우)",
                                     ["없음"] + [d['name'] for d in depts])

            if st.form_submit_button("➕ 생성", type="primary"):
                if not username or not display_name:
                    st.error("아이디와 표시명은 필수입니다.")
                else:
                    dept_id = next((d['id'] for d in depts if d['name'] == dept_name), None)
                    ok, msg = create_user(username, password, role, display_name, dept_id)
                    if ok:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)


# ═══════════════════════════════════════════════════════════════
#  MAIN ROUTER
# ═══════════════════════════════════════════════════════════════

def main():
    init_db()

    if not is_logged_in():
        page_login()
        return

    if current_user().get('must_change_pw'):
        page_change_password()
        return

    render_sidebar()

    # 기본 페이지 설정
    if 'page' not in st.session_state:
        st.session_state['page'] = 'dashboard' if is_super_admin() else 'dept_attendance'

    page = st.session_state['page']

    page_map = {
        'dashboard': page_dashboard,
        'dept_attendance': page_dept_attendance,
        'history': page_history,
        'departments': page_departments,
        'employees': page_employees,
        'salary': page_salary,
        'users': page_users,
    }

    # 권한 체크
    if not is_super_admin() and page in ['dashboard', 'departments', 'employees', 'salary', 'users']:
        st.warning("접근 권한이 없습니다.")
        st.session_state['page'] = 'dept_attendance'
        st.rerun()
        return

    render_fn = page_map.get(page, page_dashboard)
    render_fn()


if __name__ == "__main__":
    main()
