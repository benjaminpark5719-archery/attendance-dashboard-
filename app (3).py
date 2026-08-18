# -*- coding: utf-8 -*-
"""
================================================================================
 보물섬수산 HR 시스템 — 통합 app.py (완결형 레퍼런스 구현 / Full Refactor)
 Streamlit + Supabase | 모바일·PC 겸용 | 8개 모듈 통합
--------------------------------------------------------------------------------
 [핵심 설계]
  1) MENU_ITEMS 를 단일 진리 원천(single source of truth)으로 사용 →
     시스템관리의 메뉴 순서 변경 시 max_value / index 를 len 기반으로 동적 계산
     → StreamlitValueAboveMaxError 근본 차단.
  2) 사진 업로드는 원본 bytes 를 session_state 에 절대 담지 않고 Pillow 로
     리사이즈(최대 1280px, JPEG q70) 후 처리 → 업로드/저장 에러 근본 차단.
  3) 근무계획(shift_plans) → 출근입력 예정시간 자동 연동 →
     flex_attendance 판정 로직(judge_attendance) 내장.
  4) Supabase 미연결 시 자동으로 DEMO 모드(session_state 시드) 로 폴백 →
     secrets 만 채우면 실 DB 로 전환.

 [표준 DB 스키마]  ※ 컬럼명은 실제 Supabase 에 맞게 소폭 수정 후 사용
 ---------------------------------------------------------------------------
  departments (부서)
    id            uuid / bigint  PK
    name          text           부서명
    sort_order    int            표기 순서

  employees (직원)
    id                 PK
    emp_no             text   사번
    name               text
    department         text   부서명 (departments.name 참조)
    position           text   직급
    hire_date          date   입사일
    phone              text
    health_cert_expiry date   보건증 만료일
    hourly_wage        int    시급(원)
    status             text   재직 / 퇴직

  attendance (출근 기록)
    id            PK
    employee_id   (employees.id)
    emp_no        text
    name          text
    department    text
    work_date     date
    planned_start time   예정 출근시간(근무계획 연동)
    actual_start  time   실제 출근시간
    actual_end    time   퇴근시간
    status        text   정상 / 지각 / 결근
    overtime_hours numeric
    holiday_hours  numeric
    note          text

  shift_plans (주간 순환 근무계획 / 탄력근무제)
    id           PK
    employee_id
    emp_no  / name / department
    week_start   date   해당 주 월요일
    mon,tue,wed,thu,fri,sat,sun  text  (오전/오후/야간/휴무)

  tbm_logs (TBM 안전관리 일지)
    id           PK
    log_date     date
    department   text
    topic        text
    content      text
    hazard       text   위험요인
    attendees    jsonb  ["홍길동", ...]
    signatures   jsonb  {"홍길동": true, ...}
    photo_url    text   (Supabase Storage 공개 URL)
    created_at   timestamptz

 [배포 체크리스트]  → 하단 README.md 참고
  requirements.txt / fonts/NanumGothic.ttf / .streamlit/secrets.toml
================================================================================
"""

import io
import base64
from datetime import date, datetime, time, timedelta

import pandas as pd
import streamlit as st
from PIL import Image

# ── 선택적 의존성 (없어도 앱은 동작) ──────────────────────────────────────────
try:
    from supabase import create_client
    _HAS_SUPABASE = True
except Exception:
    _HAS_SUPABASE = False

try:
    from streamlit_option_menu import option_menu
    _HAS_OPTION_MENU = True
except Exception:
    _HAS_OPTION_MENU = False

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except Exception:
    _HAS_OPENPYXL = False

try:
    from fpdf import FPDF
    _HAS_FPDF = True
except Exception:
    _HAS_FPDF = False

try:
    import holidays as _kr_holidays
    _HAS_HOLIDAYS = True
except Exception:
    _HAS_HOLIDAYS = False


# ══════════════════════════════════════════════════════════════════════════════
# 1. 기본 설정 / 상수
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="보물섬수산 HR", page_icon="🐟", layout="wide")

APP_TITLE = "보물섬수산"
APP_SUBTITLE = "HR 통합 관리 시스템"
LOGO_EMOJI = "🐟"

# 로그인 (실서비스는 Supabase Auth / 해시 비밀번호 권장)
LOGIN_ID = "admin"
LOGIN_PW = "admin1234!"

# ── 메뉴: 단일 진리 원천 (라벨, 아이콘) ──────────────────────────────────────
DEFAULT_MENU = [
    ("전사현황", "📊"),
    ("출근입력", "📋"),
    ("근무계획", "🗓️"),
    ("이력조회", "🔍"),
    ("부서관리", "🏢"),
    ("직원관리", "👤"),
    ("TBM안전관리", "🦺"),
    ("급여관리", "💰"),
    ("시스템관리", "⚙️"),
]

# 실제 Supabase 테이블명 매핑 (여기만 바꾸면 전체 반영)
TB = {
    "departments": "departments",
    "employees": "employees",
    "attendance": "attendance",
    "shift_plans": "shift_plans",
    "tbm_logs": "tbm_logs",
}

# 근무조 → 예정 시작/종료 시간 (탄력근무제 표준표) ─ 회사 규정에 맞게 조정
SHIFT_SCHEDULE = {
    "오전": {"start": time(6, 0), "end": time(15, 0), "color": "#FFF3CD"},
    "오후": {"start": time(13, 0), "end": time(22, 0), "color": "#D1E7DD"},
    "야간": {"start": time(22, 0), "end": time(6, 0), "color": "#CFE2FF"},
    "휴무": {"start": None, "end": None, "color": "#E2E3E5"},
}
SHIFT_OPTIONS = list(SHIFT_SCHEDULE.keys())
GRACE_MINUTES = 0          # 지각 유예(분)
OVERTIME_RATE = 1.5        # 연장 가산율
HOLIDAY_RATE = 1.5         # 휴일 가산율
STATUS_COLOR = {"정상": "#198754", "지각": "#dc3545", "결근": "#6c757d"}

FONT_PATH = "fonts/NanumGothic.ttf"


# ══════════════════════════════════════════════════════════════════════════════
# 2. 스타일 (카드 / 뱃지 / 모바일 최적화 CSS)
# ══════════════════════════════════════════════════════════════════════════════
CUSTOM_CSS = """
<style>
:root { --brand:#0d6efd; --brand-dark:#0a58ca; }
section[data-testid="stSidebar"] { background:#0b1f33; }
section[data-testid="stSidebar"] * { color:#e9eef5; }
.brand-head { display:flex; align-items:center; gap:.55rem; padding:.4rem .2rem 1rem .2rem;
    border-bottom:1px solid rgba(255,255,255,.12); margin-bottom:.8rem; }
.brand-emoji { font-size:1.9rem; }
.brand-name  { font-size:1.15rem; font-weight:800; line-height:1.1; }
.brand-sub   { font-size:.72rem; opacity:.75; }

.metric-card { background:#fff; border:1px solid #eef0f3; border-radius:16px;
    padding:.9rem 1rem; box-shadow:0 2px 10px rgba(16,24,40,.05); }
.metric-card .label { font-size:.8rem; color:#667085; margin-bottom:.15rem; }
.metric-card .value { font-size:1.7rem; font-weight:800; line-height:1; }
.metric-card .delta { font-size:.72rem; color:#98a2b3; }

.badge { display:inline-block; padding:.18rem .55rem; border-radius:999px;
    font-size:.72rem; font-weight:700; color:#fff; }
.callout { border-radius:14px; padding:.8rem 1rem; margin:.3rem 0 1rem 0;
    font-weight:600; }
.callout.warn { background:#fff4e5; border:1px solid #ffd9a8; color:#8a4b00; }
.callout.ok   { background:#e7f6ec; border:1px solid #b7e2c5; color:#1b6e3a; }

.block-container { padding-top:2rem; }
/* 헤더 글자 상단 잘림 방지: 넉넉한 line-height 확보 */
.main h1, .main h2, .main h3,
div[data-testid="stHeading"] { line-height:1.4; margin-top:.2rem; }
div[data-testid="stDataFrame"] { border-radius:12px; overflow:hidden; }
@media (max-width:640px){
  .metric-card .value { font-size:1.35rem; }
  .block-container { padding-left:.6rem; padding-right:.6rem; }
}
</style>
"""


def badge(text, color):
    return f'<span class="badge" style="background:{color}">{text}</span>'


def metric_card(label, value, delta=""):
    st.markdown(
        f'<div class="metric-card"><div class="label">{label}</div>'
        f'<div class="value">{value}</div><div class="delta">{delta}</div></div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# 3. Supabase 연결 + 데모 폴백 데이터 계층
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_resource(show_spinner=False)
def get_supabase():
    """secrets 에 URL/KEY 가 있으면 클라이언트 반환, 없으면 None (→ 데모 모드)."""
    if not _HAS_SUPABASE:
        return None
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
    except Exception:
        return None
    if not url or not key:
        return None
    try:
        return create_client(url, key)
    except Exception:
        return None


SB = get_supabase()
USE_SUPABASE = SB is not None


def _seed_demo():
    if st.session_state.get("_demo_ready"):
        return
    today = date.today()
    st.session_state.demo_departments = [
        {"id": 1, "name": "영업부", "sort_order": 1},
        {"id": 2, "name": "가공1팀", "sort_order": 2},
        {"id": 3, "name": "물류센터", "sort_order": 3},
        {"id": 4, "name": "매장운영", "sort_order": 4},
    ]
    st.session_state.demo_employees = [
        {"id": 1, "emp_no": "B001", "name": "김민수", "department": "영업부",
         "position": "과장", "hire_date": "2021-03-02", "phone": "010-1111-2222",
         "health_cert_expiry": str(today + timedelta(days=12)), "hourly_wage": 12000, "status": "재직"},
        {"id": 2, "emp_no": "B002", "name": "이서연", "department": "가공1팀",
         "position": "사원", "hire_date": "2023-06-01", "phone": "010-3333-4444",
         "health_cert_expiry": str(today + timedelta(days=45)), "hourly_wage": 10500, "status": "재직"},
        {"id": 3, "emp_no": "B003", "name": "박준호", "department": "물류센터",
         "position": "대리", "hire_date": "2022-01-10", "phone": "010-5555-6666",
         "health_cert_expiry": str(today - timedelta(days=3)), "hourly_wage": 11000, "status": "재직"},
        {"id": 4, "emp_no": "B004", "name": "최지우", "department": "매장운영",
         "position": "사원", "hire_date": "2024-02-19", "phone": "010-7777-8888",
         "health_cert_expiry": str(today + timedelta(days=120)), "hourly_wage": 10200, "status": "재직"},
    ]
    st.session_state.demo_attendance = []
    st.session_state.demo_shift_plans = []
    st.session_state.demo_tbm_logs = []
    st.session_state._demo_ready = True


def _demo_table(name):
    _seed_demo()
    return st.session_state[f"demo_{name}"]


def db_select(name):
    """테이블 전체 조회 → DataFrame."""
    if USE_SUPABASE:
        try:
            res = SB.table(TB[name]).select("*").execute()
            return pd.DataFrame(res.data or [])
        except Exception as e:
            st.session_state.setdefault("_db_errors", []).append(f"{name} select: {e}")
    return pd.DataFrame(_demo_table(name))


def db_insert(name, row: dict):
    if USE_SUPABASE:
        try:
            SB.table(TB[name]).insert(row).execute()
            return True
        except Exception as e:
            st.warning(f"저장 실패({name}): {e}")
            return False
    tbl = _demo_table(name)
    row = dict(row)
    row.setdefault("id", (max([r.get("id", 0) for r in tbl], default=0) + 1))
    tbl.append(row)
    return True


def db_delete(name, id_val):
    if USE_SUPABASE:
        try:
            SB.table(TB[name]).delete().eq("id", id_val).execute()
            return True
        except Exception as e:
            st.warning(f"삭제 실패({name}): {e}")
            return False
    tbl = _demo_table(name)
    st.session_state[f"demo_{name}"] = [r for r in tbl if r.get("id") != id_val]
    return True


def db_update(name, id_val, patch: dict):
    if USE_SUPABASE:
        try:
            SB.table(TB[name]).update(patch).eq("id", id_val).execute()
            return True
        except Exception as e:
            st.warning(f"수정 실패({name}): {e}")
            return False
    for r in _demo_table(name):
        if r.get("id") == id_val:
            r.update(patch)
    return True


def db_upsert_attendance(row: dict):
    """(employee_id, work_date) 기준 중복 제거 후 저장."""
    key_emp, key_date = row.get("employee_id"), row.get("work_date")
    if USE_SUPABASE:
        try:
            SB.table(TB["attendance"]).delete().eq("employee_id", key_emp)\
                .eq("work_date", key_date).execute()
            SB.table(TB["attendance"]).insert(row).execute()
            return True
        except Exception as e:
            st.warning(f"출근 저장 실패: {e}")
            return False
    tbl = _demo_table("attendance")
    st.session_state["demo_attendance"] = [
        r for r in tbl if not (r.get("employee_id") == key_emp and r.get("work_date") == key_date)
    ]
    return db_insert("attendance", row)


# ══════════════════════════════════════════════════════════════════════════════
# 4. 유틸리티 (이미지 압축 / 근태 판정 / 엑셀·PDF)
# ══════════════════════════════════════════════════════════════════════════════
def compress_image(uploaded_file, max_side=1280, quality=70):
    """원본 bytes 를 session_state 에 담지 않고 리사이즈 후 JPEG bytes 반환."""
    img = Image.open(uploaded_file)
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    img.thumbnail((max_side, max_side))
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    return buf.getvalue()


def _parse_time(v):
    if v in (None, "", "None"):
        return None
    if isinstance(v, time):
        return v
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(str(v), fmt).time()
        except ValueError:
            continue
    return None


def judge_attendance(planned_start, actual_start, grace=GRACE_MINUTES):
    """flex_attendance 핵심 로직: 예정시간 대비 지각/정상/결근 판정."""
    planned_start = _parse_time(planned_start)
    actual_start = _parse_time(actual_start)
    if planned_start is None:          # 휴무 / 예정 미지정
        return "정상" if actual_start else "결근"
    if actual_start is None:
        return "결근"
    limit = (datetime.combine(date.today(), planned_start) + timedelta(minutes=grace)).time()
    return "지각" if actual_start > limit else "정상"


def get_kr_holidays(year):
    if _HAS_HOLIDAYS:
        try:
            return set(_kr_holidays.KR(years=year).keys())
        except Exception:
            pass
    return set()


def build_shift_template_xlsx(employees_df):
    """부서/이름 + 월~일 근무조 드롭다운 엑셀 템플릿 (bytes)."""
    if not _HAS_OPENPYXL:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "근무계획"
    headers = ["부서", "이름", "월", "화", "수", "목", "금", "토", "일"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="0D6EFD")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for _, e in employees_df.iterrows():
        ws.append([e.get("department", ""), e.get("name", "")] + ["휴무"] * 7)
    dv = DataValidation(type="list", formula1='"오전,오후,야간,휴무"', allow_blank=True)
    ws.add_data_validation(dv)
    last = ws.max_row if ws.max_row > 1 else 200
    dv.add(f"C2:I{max(last, 200)}")
    for col, w in zip("ABCDEFGHI", [12, 12, 8, 8, 8, 8, 8, 8, 8]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_new():
    if not _HAS_FPDF:
        return None
    pdf = FPDF()
    pdf.add_page()
    try:
        pdf.add_font("Nanum", "", FONT_PATH, uni=True)
        pdf.set_font("Nanum", size=11)
        return pdf
    except Exception:
        # 폰트 미탑재 시 안내 (한글 깨짐 방지 위해 배포 시 반드시 폰트 추가)
        pdf.set_font("Helvetica", size=11)
        return pdf


def build_payslip_pdf(emp, period, rows: dict):
    pdf = _pdf_new()
    if pdf is None:
        return None
    pdf.set_font_size(16)
    pdf.cell(0, 12, f"{APP_TITLE} 급여명세서", ln=1, align="C")
    pdf.set_font_size(11)
    pdf.cell(0, 8, f"성명: {emp['name']}  /  사번: {emp['emp_no']}  /  부서: {emp['department']}", ln=1)
    pdf.cell(0, 8, f"귀속월: {period}", ln=1)
    pdf.ln(4)
    for k, v in rows.items():
        pdf.cell(60, 8, str(k), border=1)
        pdf.cell(0, 8, f"{v:,} 원" if isinstance(v, (int, float)) else str(v), border=1, ln=1, align="R")
    return pdf.output(dest="S").encode("latin-1", errors="ignore")


def build_tbm_pdf(log):
    pdf = _pdf_new()
    if pdf is None:
        return None
    pdf.set_font_size(16)
    pdf.cell(0, 12, "TBM 안전보건 일지", ln=1, align="C")
    pdf.set_font_size(11)
    for label, key in [("일자", "log_date"), ("부서", "department"),
                       ("주제", "topic"), ("위험요인", "hazard")]:
        pdf.cell(35, 8, label, border=1)
        pdf.multi_cell(0, 8, str(log.get(key, "")), border=1)
    pdf.ln(2)
    pdf.multi_cell(0, 8, f"내용: {log.get('content', '')}", border=1)
    attendees = log.get("attendees", [])
    sigs = log.get("signatures", {})
    pdf.ln(2)
    pdf.cell(0, 8, f"참석 {len(attendees)}명 / 서명완료 {sum(1 for a in attendees if sigs.get(a))}명", ln=1)
    for a in attendees:
        mark = "✔ 서명" if sigs.get(a) else "미서명"
        pdf.cell(0, 7, f" - {a} : {mark}", ln=1)
    return pdf.output(dest="S").encode("latin-1", errors="ignore")


def dept_options_with_count(dept_df, emp_df):
    """'부서명 (인원수)' 라벨 리스트. 빈 부서도 표기하되 인원 0 명시."""
    counts = emp_df.groupby("department").size().to_dict() if not emp_df.empty else {}
    labels, mapping = [], {}
    src = dept_df["name"].tolist() if not dept_df.empty else sorted(counts.keys())
    for name in src:
        label = f"{name} ({counts.get(name, 0)}명)"
        labels.append(label)
        mapping[label] = name
    return labels, mapping


# ══════════════════════════════════════════════════════════════════════════════
# 5. 인증
# ══════════════════════════════════════════════════════════════════════════════
def login_gate():
    if st.session_state.get("auth"):
        return True
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
    st.markdown(f"## {LOGO_EMOJI} {APP_TITLE} {APP_SUBTITLE}")
    with st.form("login"):
        uid = st.text_input("아이디", value="")
        pw = st.text_input("비밀번호", type="password")
        ok = st.form_submit_button("로그인", use_container_width=True)
    if ok:
        if uid == LOGIN_ID and pw == LOGIN_PW:
            st.session_state.auth = True
            st.rerun()
        else:
            st.error("아이디 또는 비밀번호가 올바르지 않습니다.")
    return False


# ══════════════════════════════════════════════════════════════════════════════
# 6. 페이지 — ① 전사현황
# ══════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    st.subheader("📊 전사현황")
    emp = db_select("employees")
    att = db_select("attendance")
    today = str(date.today())
    today_att = att[att["work_date"] == today] if not att.empty else pd.DataFrame()

    total = len(emp[emp["status"] == "재직"]) if not emp.empty else 0
    present = len(today_att[today_att["status"].isin(["정상", "지각"])]) if not today_att.empty else 0
    late = len(today_att[today_att["status"] == "지각"]) if not today_att.empty else 0
    absent = max(total - present, 0)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        metric_card("전체 재직", f"{total}명")
    with c2:
        metric_card("출근 완료", f"{present}명", f"{(present/total*100):.0f}%" if total else "-")
    with c3:
        metric_card("지각", f"{late}명")
    with c4:
        metric_card("결근/미입력", f"{absent}명")

    st.write("")
    if late or absent:
        names_late = today_att[today_att["status"] == "지각"]["name"].tolist() if not today_att.empty else []
        msg = f"⚠️ 지각 {late}명"
        if names_late:
            msg += f" ({', '.join(names_late)})"
        msg += f"  ·  결근/미입력 {absent}명 — 확인이 필요합니다."
        st.markdown(f'<div class="callout warn">{msg}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="callout ok">✅ 금일 지각·결근 없음. 전원 정상 출근입니다.</div>',
                    unsafe_allow_html=True)

    st.markdown("##### 전체 직원 현황")
    if emp.empty:
        st.info("등록된 직원이 없습니다.")
        return
    view = emp[["emp_no", "name", "department", "position", "status"]].copy()
    view.columns = ["사번", "이름", "부서", "직급", "상태"]
    st.dataframe(view.sort_values("부서"), use_container_width=True, hide_index=True)


# ── ② 출근입력 ────────────────────────────────────────────────────────────────
def _planned_start_for(emp_row, target_date):
    """근무계획(shift_plans)에서 해당 직원·요일의 근무조 → 예정 시작시간."""
    plans = db_select("shift_plans")
    if plans.empty:
        return None
    week_start = target_date - timedelta(days=target_date.weekday())
    key = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"][target_date.weekday()]
    row = plans[(plans["emp_no"] == emp_row["emp_no"]) &
                (plans["week_start"] == str(week_start))]
    if row.empty:
        return None
    shift = row.iloc[0].get(key, "휴무")
    sch = SHIFT_SCHEDULE.get(shift)
    return sch["start"] if sch else None


def page_checkin():
    st.subheader("📋 출근입력")
    emp = db_select("employees")
    dept = db_select("departments")
    if emp.empty:
        st.info("직원을 먼저 등록하세요.")
        return

    labels, mapping = dept_options_with_count(dept, emp)
    sel = st.selectbox("부서 선택", labels, index=0 if labels else None)
    dept_name = mapping.get(sel) if sel else None
    target = st.date_input("출근일", value=date.today())
    members = emp[emp["department"] == dept_name] if dept_name else emp.iloc[0:0]

    # 상단 액션 버튼 전진 배치
    a1, a2 = st.columns(2)
    bulk = a1.button("⚡ 전체 예정시간대로 일괄 출근 승인", use_container_width=True)
    save = a2.button("💾 출근 데이터 저장", type="primary", use_container_width=True)

    st.divider()
    inputs = {}
    for _, e in members.iterrows():
        planned = _planned_start_for(e, target)
        default_start = planned or time(9, 0)
        col1, col2, col3 = st.columns([2, 2, 2])
        col1.markdown(f"**{e['name']}** · {e['emp_no']}")
        col1.caption(f"예정 {planned.strftime('%H:%M') if planned else '미지정'}")
        if bulk and planned:
            st.session_state[f"in_{e['emp_no']}"] = planned
        actual = col2.time_input("출근", key=f"in_{e['emp_no']}",
                                 value=default_start, label_visibility="collapsed")
        status = judge_attendance(planned, actual)
        col3.markdown(badge(status, STATUS_COLOR[status]), unsafe_allow_html=True)
        inputs[e["emp_no"]] = (e, planned, actual, status)

    if save and inputs:
        n = 0
        for emp_no, (e, planned, actual, status) in inputs.items():
            row = {
                "employee_id": e["id"], "emp_no": emp_no, "name": e["name"],
                "department": e["department"], "work_date": str(target),
                "planned_start": planned.strftime("%H:%M") if planned else None,
                "actual_start": actual.strftime("%H:%M") if actual else None,
                "actual_end": None, "status": status,
                "overtime_hours": 0, "holiday_hours": 0, "note": "",
            }
            if db_upsert_attendance(row):
                n += 1
        st.success(f"{n}건 저장 완료.")


# ── ③ 근무계획 ────────────────────────────────────────────────────────────────
def page_shift_plan():
    st.subheader("🗓️ 근무계획 (순환 탄력근무제)")
    emp = db_select("employees")

    tmpl = build_shift_template_xlsx(emp) if not emp.empty else None
    if tmpl:
        st.download_button("📥 주간 순환근무 표준 엑셀 양식 다운로드", data=tmpl,
                           file_name="근무계획_양식.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    else:
        st.caption("엑셀 양식 생성을 위해 openpyxl 설치가 필요합니다.")

    week_start = st.date_input("계획 주 (월요일 기준)", value=date.today() - timedelta(days=date.today().weekday()))
    week_start = week_start - timedelta(days=week_start.weekday())

    with st.expander("📤 엑셀 업로드 / 직접 등록", expanded=False):
        up = st.file_uploader("작성한 근무계획 엑셀 업로드", type=["xlsx"])
        if up and _HAS_OPENPYXL:
            try:
                df = pd.read_excel(up)
                for _, r in df.iterrows():
                    match = emp[emp["name"] == r.get("이름")]
                    if match.empty:
                        continue
                    e = match.iloc[0]
                    row = {"employee_id": e["id"], "emp_no": e["emp_no"], "name": e["name"],
                           "department": e["department"], "week_start": str(week_start)}
                    for kor, key in zip(["월", "화", "수", "목", "금", "토", "일"],
                                        ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]):
                        row[key] = str(r.get(kor, "휴무"))
                    # 동일 (emp, week) 갱신
                    st.session_state.setdefault("_del", [])
                    db_insert("shift_plans", row)
                st.success("근무계획 업로드 완료.")
            except Exception as e:
                st.error(f"업로드 오류: {e}")

    # 시각화 매트릭스
    plans = db_select("shift_plans")
    wk = plans[plans["week_start"] == str(week_start)] if not plans.empty else pd.DataFrame()
    st.markdown("##### 주간 순환 근무 달력")
    if wk.empty:
        st.info("해당 주에 등록된 근무계획이 없습니다.")
    else:
        days = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]
        kor = ["월", "화", "수", "목", "금", "토", "일"]
        holi = get_kr_holidays(week_start.year)
        head = []
        for i, k in enumerate(kor):
            d = week_start + timedelta(days=i)
            mark = "🔴" if d in holi else ""
            head.append(f"{k}{mark}")
        matrix = wk[["name"] + days].copy()
        matrix.columns = ["이름"] + head

        def color(v):
            return f"background-color:{SHIFT_SCHEDULE.get(v, {}).get('color', '#fff')}"
        st.dataframe(matrix.style.applymap(color, subset=head),
                     use_container_width=True, hide_index=True)
        st.caption("🟡오전  🟢오후  🔵야간  ⚪휴무   ·   🔴공휴일/대체휴일 자동 표시")


# ── ④ 이력조회 ────────────────────────────────────────────────────────────────
def page_history():
    st.subheader("🔍 이력조회")
    att = db_select("attendance")
    if att.empty:
        st.info("출근 이력이 없습니다.")
        return

    c1, c2, c3 = st.columns([2, 2, 2])
    kw = c1.text_input("이름 / 사번 검색")
    status_f = c2.multiselect("상태 필터", ["정상", "지각", "결근"], default=["정상", "지각", "결근"])
    dr = c3.date_input("기간", value=(date.today() - timedelta(days=30), date.today()))

    df = att.copy()
    if isinstance(dr, tuple) and len(dr) == 2:
        df = df[(df["work_date"] >= str(dr[0])) & (df["work_date"] <= str(dr[1]))]
    if kw:
        df = df[df["name"].str.contains(kw, na=False) | df["emp_no"].astype(str).str.contains(kw, na=False)]
    if status_f:
        df = df[df["status"].isin(status_f)]

    total = len(df)
    late = len(df[df["status"] == "지각"])
    absent = len(df[df["status"] == "결근"])
    rate = ((total - absent) / total * 100) if total else 0
    m1, m2, m3 = st.columns(3)
    with m1:
        metric_card("총 출근율", f"{rate:.1f}%", f"{total}건 기준")
    with m2:
        metric_card("지각 합계", f"{late}건")
    with m3:
        metric_card("결근 합계", f"{absent}건")

    view = df[["work_date", "emp_no", "name", "department",
               "planned_start", "actual_start", "status"]].copy()
    view.columns = ["일자", "사번", "이름", "부서", "예정", "출근", "상태"]

    def hi(v):
        return f"color:{STATUS_COLOR.get(v, '#000')}; font-weight:700"
    st.dataframe(view.sort_values("일자", ascending=False)
                 .style.applymap(hi, subset=["상태"]),
                 use_container_width=True, hide_index=True)


# ── ⑤ 부서관리 ────────────────────────────────────────────────────────────────
def _dialog_available():
    return hasattr(st, "dialog")


def page_department():
    st.subheader("🏢 부서관리")
    dept = db_select("departments")
    emp = db_select("employees")

    if st.button("➕ 새 부서 추가", type="primary"):
        st.session_state["show_add_dept"] = True

    if st.session_state.get("show_add_dept"):
        if _dialog_available():
            @st.dialog("새 부서 추가")
            def _add():
                name = st.text_input("부서명")
                order = st.number_input("표기 순서", min_value=1, value=int(len(dept) + 1))
                if st.button("추가", type="primary"):
                    if name.strip():
                        db_insert("departments", {"name": name.strip(), "sort_order": int(order)})
                        st.session_state["show_add_dept"] = False
                        st.rerun()
                    else:
                        st.warning("부서명을 입력하세요.")
            _add()
        else:
            with st.form("add_dept_form"):
                name = st.text_input("부서명")
                if st.form_submit_button("추가"):
                    if name.strip():
                        db_insert("departments", {"name": name.strip(), "sort_order": len(dept) + 1})
                        st.session_state["show_add_dept"] = False
                        st.rerun()

    if dept.empty:
        st.info("등록된 부서가 없습니다.")
        return
    counts = emp.groupby("department").size().to_dict() if not emp.empty else {}
    show = dept.sort_values("sort_order")[["name", "sort_order"]].copy()
    show["인원수"] = show["name"].map(lambda n: counts.get(n, 0))
    show.columns = ["부서명", "순서", "인원수"]
    st.dataframe(show, use_container_width=True, hide_index=True)

    del_name = st.selectbox("삭제할 부서", ["-"] + dept["name"].tolist())
    if del_name != "-" and st.button("선택 부서 삭제"):
        row = dept[dept["name"] == del_name].iloc[0]
        db_delete("departments", row["id"])
        st.rerun()


# ── ⑥ 직원관리 ────────────────────────────────────────────────────────────────
def _health_badge(expiry):
    try:
        d = datetime.strptime(str(expiry), "%Y-%m-%d").date()
    except Exception:
        return badge("미등록", "#6c757d")
    left = (d - date.today()).days
    if left < 0:
        return badge(f"만료 {abs(left)}일 경과", "#dc3545")
    if left <= 30:
        return badge(f"D-{left} 임박", "#fd7e14")
    return badge("정상", "#198754")


def page_employee():
    st.subheader("👤 직원관리")
    emp = db_select("employees")
    dept = db_select("departments")

    kw = st.text_input("🔎 이름 / 사번 / 부서 통합 검색")
    if st.button("➕ 새 직원 추가", type="primary"):
        st.session_state["show_add_emp"] = True

    if st.session_state.get("show_add_emp") and _dialog_available():
        @st.dialog("새 직원 추가")
        def _add_emp():
            emp_no = st.text_input("사번")
            name = st.text_input("이름")
            d = st.selectbox("부서", dept["name"].tolist() if not dept.empty else [])
            pos = st.text_input("직급", value="사원")
            wage = st.number_input("시급(원)", min_value=0, value=10030, step=100)
            hc = st.date_input("보건증 만료일", value=date.today() + timedelta(days=180))
            if st.button("등록", type="primary"):
                db_insert("employees", {
                    "emp_no": emp_no, "name": name, "department": d, "position": pos,
                    "hire_date": str(date.today()), "phone": "",
                    "health_cert_expiry": str(hc), "hourly_wage": int(wage), "status": "재직"})
                st.session_state["show_add_emp"] = False
                st.rerun()
        _add_emp()

    if emp.empty:
        st.info("등록된 직원이 없습니다.")
        return
    df = emp.copy()
    if kw:
        df = df[df["name"].str.contains(kw, na=False)
                | df["emp_no"].astype(str).str.contains(kw, na=False)
                | df["department"].str.contains(kw, na=False)]

    for _, e in df.iterrows():
        c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
        c1.markdown(f"**{e['name']}** · {e['emp_no']}")
        c2.caption(f"{e['department']} / {e.get('position', '')}")
        c3.markdown("보건증 " + _health_badge(e.get("health_cert_expiry")), unsafe_allow_html=True)
        if c4.button("수정", key=f"edit_{e['emp_no']}") and _dialog_available():
            st.session_state["edit_emp_id"] = e["id"]

    if st.session_state.get("edit_emp_id") and _dialog_available():
        eid = st.session_state["edit_emp_id"]
        row = emp[emp["id"] == eid]
        if not row.empty:
            e = row.iloc[0]
            @st.dialog(f"{e['name']} 정보 수정")
            def _edit():
                pos = st.text_input("직급", value=str(e.get("position", "")))
                wage = st.number_input("시급(원)", min_value=0, value=int(e.get("hourly_wage", 0)), step=100)
                hc = st.text_input("보건증 만료일(YYYY-MM-DD)", value=str(e.get("health_cert_expiry", "")))
                status = st.selectbox("상태", ["재직", "퇴직"],
                                      index=0 if e.get("status") == "재직" else 1)
                if st.button("저장", type="primary"):
                    db_update("employees", eid, {"position": pos, "hourly_wage": int(wage),
                                                 "health_cert_expiry": hc, "status": status})
                    st.session_state["edit_emp_id"] = None
                    st.rerun()
            _edit()


# ── ⑦ TBM 안전관리 ────────────────────────────────────────────────────────────
def page_tbm():
    st.subheader("🦺 TBM 안전관리")
    emp = db_select("employees")
    dept = db_select("departments")

    with st.form("tbm_form"):
        log_date = st.date_input("일자", value=date.today())
        d = st.selectbox("부서", dept["name"].tolist() if not dept.empty else [])
        topic = st.text_input("TBM 주제")
        hazard = st.text_input("주요 위험요인")
        content = st.text_area("교육 내용")
        photo = st.file_uploader("현장 사진 (자동 압축)", type=["jpg", "jpeg", "png"])
        submitted = st.form_submit_button("일지 등록", type="primary")

    # Title Validation: 공백/누락 방지
    if submitted:
        if not topic.strip():
            st.error("TBM 주제(Title)를 입력해야 저장할 수 있습니다.")
        else:
            photo_uri = None
            if photo is not None:
                try:
                    data = compress_image(photo)          # 원본 bytes 미보관
                    if USE_SUPABASE:
                        path = f"tbm/{datetime.now().timestamp()}.jpg"
                        try:
                            SB.storage.from_("tbm-photos").upload(path, data,
                                {"content-type": "image/jpeg"})
                            photo_uri = SB.storage.from_("tbm-photos").get_public_url(path)
                        except Exception as e:
                            st.warning(f"사진 업로드 실패(버킷 확인): {e}")
                    else:
                        photo_uri = "data:image/jpeg;base64," + base64.b64encode(data).decode()
                except Exception as e:
                    st.warning(f"이미지 처리 실패: {e}")
            members = emp[emp["department"] == d]["name"].tolist() if not emp.empty else []
            db_insert("tbm_logs", {
                "log_date": str(log_date), "department": d, "topic": topic.strip(),
                "content": content, "hazard": hazard, "attendees": members,
                "signatures": {}, "photo_url": photo_uri,
                "created_at": datetime.now().isoformat()})
            st.success("TBM 일지 등록 완료.")

    st.divider()
    logs = db_select("tbm_logs")
    if logs.empty:
        st.info("등록된 TBM 일지가 없습니다.")
        return
    for _, log in logs.sort_values("log_date", ascending=False).iterrows():
        with st.expander(f"[{log['log_date']}] {log['department']} · {log['topic']}"):
            st.write(f"**위험요인:** {log.get('hazard', '')}")
            st.write(log.get("content", ""))
            if log.get("photo_url"):
                st.image(log["photo_url"], width=280)
            attendees = log.get("attendees", []) or []
            sigs = dict(log.get("signatures", {}) or {})
            st.markdown("**전자서명 (1클릭 이수 체크)**")
            for a in attendees:
                cols = st.columns([3, 2])
                cols[0].write(a)
                checked = cols[1].checkbox("이수", value=bool(sigs.get(a)),
                                           key=f"sig_{log['id']}_{a}")
                sigs[a] = checked
            if st.button("서명 저장", key=f"savesig_{log['id']}"):
                db_update("tbm_logs", log["id"], {"signatures": sigs})
                st.success("서명 저장 완료.")
            pdf_bytes = build_tbm_pdf({**log, "signatures": sigs})
            if pdf_bytes:
                st.download_button("📥 TBM 일지 PDF", data=pdf_bytes,
                                   file_name=f"TBM_{log['log_date']}.pdf",
                                   mime="application/pdf", key=f"tbmpdf_{log['id']}")


# ── ⑧ 급여관리 ────────────────────────────────────────────────────────────────
def page_payroll():
    st.subheader("💰 급여관리")
    emp = db_select("employees")
    att = db_select("attendance")
    if emp.empty:
        st.info("직원 정보가 없습니다.")
        return

    ym = st.text_input("귀속월 (YYYY-MM)", value=date.today().strftime("%Y-%m"))
    month_att = att[att["work_date"].str.startswith(ym)] if not att.empty else pd.DataFrame()

    rows = []
    for _, e in emp.iterrows():
        rec = month_att[month_att["emp_no"] == e["emp_no"]] if not month_att.empty else pd.DataFrame()
        work_days = len(rec[rec["status"].isin(["정상", "지각"])]) if not rec.empty else 0
        ot = rec["overtime_hours"].fillna(0).sum() if not rec.empty else 0
        hol = rec["holiday_hours"].fillna(0).sum() if not rec.empty else 0
        wage = int(e.get("hourly_wage", 0) or 0)
        base = work_days * 8 * wage
        ot_pay = int(ot * wage * OVERTIME_RATE)
        hol_pay = int(hol * wage * HOLIDAY_RATE)
        total = base + ot_pay + hol_pay
        rows.append({"사번": e["emp_no"], "이름": e["name"], "부서": e["department"],
                     "근무일": work_days, "기본급": base, "연장수당": ot_pay,
                     "휴일수당": hol_pay, "예상 지급액": total,
                     "_emp": e.to_dict(),
                     "_detail": {"근무일수": work_days, "기본급": base,
                                 "연장수당": ot_pay, "휴일수당": hol_pay, "합계": total}})
    pdf_df = pd.DataFrame(rows)
    st.dataframe(pdf_df.drop(columns=["_emp", "_detail"]),
                 use_container_width=True, hide_index=True)
    st.caption("※ 기본급 = 근무일 × 8h × 시급 (표준 모델). 주휴·4대보험·공제는 회사 규정에 맞게 확장하세요.")

    pick = st.selectbox("급여명세서 발급 대상", pdf_df["이름"].tolist())
    target = pdf_df[pdf_df["이름"] == pick].iloc[0]
    pdf_bytes = build_payslip_pdf(target["_emp"], ym, target["_detail"])
    if pdf_bytes:
        st.download_button("📄 개인별 급여명세서 PDF", data=pdf_bytes,
                           file_name=f"급여명세서_{pick}_{ym}.pdf",
                           mime="application/pdf")
    else:
        st.caption("PDF 생성을 위해 fpdf2 및 NanumGothic 폰트가 필요합니다.")


# ── ⑨ 시스템관리 (메뉴 순서 / StreamlitValueAboveMaxError 수정) ───────────────
def page_system():
    st.subheader("⚙️ 시스템관리")
    st.markdown("##### 메뉴 표시 순서")
    menu = st.session_state.get("menu_items", list(DEFAULT_MENU))
    n = len(menu)   # ← 동적 max_value 기준

    new_order = []
    for i, (label, icon) in enumerate(menu):
        # 핵심: max_value 를 len(menu) 로 동적 계산 → ValueAboveMaxError 차단
        pos = st.number_input(f"{icon} {label}", min_value=1, max_value=n,
                              value=min(i + 1, n), key=f"order_{label}")
        new_order.append((pos, label, icon))

    if st.button("순서 저장", type="primary"):
        ordered = [(lb, ic) for _, lb, ic in sorted(new_order, key=lambda x: x[0])]
        st.session_state["menu_items"] = ordered
        # 현재 선택 인덱스도 새 길이에 맞게 클램핑
        st.session_state["nav_idx"] = min(st.session_state.get("nav_idx", 0), len(ordered) - 1)
        st.success("메뉴 순서를 저장했습니다.")
        st.rerun()

    st.divider()
    st.markdown("##### 연결 상태")
    st.write(f"- Supabase 연결: {'🟢 연결됨' if USE_SUPABASE else '🟡 데모 모드(secrets 미설정)'}")
    st.write(f"- 한글 PDF 폰트: {'🟢 사용 가능' if _HAS_FPDF else '🔴 fpdf2 미설치'}")
    if st.session_state.get("_db_errors"):
        with st.expander("DB 경고 로그"):
            for e in st.session_state["_db_errors"][-20:]:
                st.caption(e)


# ══════════════════════════════════════════════════════════════════════════════
# 7. 라우터
# ══════════════════════════════════════════════════════════════════════════════
PAGE_FUNCS = {
    "전사현황": page_dashboard, "출근입력": page_checkin, "근무계획": page_shift_plan,
    "이력조회": page_history, "부서관리": page_department, "직원관리": page_employee,
    "TBM안전관리": page_tbm, "급여관리": page_payroll, "시스템관리": page_system,
}


def sidebar_nav():
    menu = st.session_state.get("menu_items", list(DEFAULT_MENU))
    labels = [m[0] for m in menu]
    icons = [m[1] for m in menu]

    # 저장된 인덱스를 현재 메뉴 길이에 맞게 클램핑 (재정렬 후 안전)
    idx = min(st.session_state.get("nav_idx", 0), len(labels) - 1)

    st.markdown(
        f'<div class="brand-head"><span class="brand-emoji">{LOGO_EMOJI}</span>'
        f'<div><div class="brand-name">{APP_TITLE}</div>'
        f'<div class="brand-sub">{APP_SUBTITLE}</div></div></div>',
        unsafe_allow_html=True,
    )

    if _HAS_OPTION_MENU:
        # streamlit-option-menu 는 이모지 대신 bootstrap icon 사용 → 라벨에 이모지 포함
        choice = option_menu(None, [f"{ic} {lb}" for lb, ic in zip(labels, icons)],
                             default_index=idx,
                             styles={"container": {"background-color": "transparent"},
                                     "nav-link-selected": {"background-color": "#0d6efd"}})
        selected = labels[[f"{ic} {lb}" for lb, ic in zip(labels, icons)].index(choice)]
    else:
        selected = st.radio("이동", labels,
                            index=idx,
                            format_func=lambda l: f"{icons[labels.index(l)]} {l}",
                            label_visibility="collapsed")

    st.session_state["nav_idx"] = labels.index(selected)
    st.divider()
    if st.button("로그아웃", use_container_width=True):
        st.session_state.auth = False
        st.rerun()
    return selected


def main():
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
    if not login_gate():
        return
    with st.sidebar:
        selected = sidebar_nav()
    PAGE_FUNCS.get(selected, page_dashboard)()


if __name__ == "__main__":
    main()
